# -*- coding: utf-8 -*-
"""
한국 국채선물(KTB Futures) 차익거래 실시간 모니터 엑셀 생성기

산출물: 국채선물_차익거래.xlsx
  - 실시간   : RTD/DDE 연결 셀 (선물가격, 바스켓 호가수익률, 조달금리)
  - 설정     : 계약사양 / 일정 / 금리 / 포지션 / 거래비용
  - 바스켓   : 최종결제기준채권 3종목 현물·선도 평가, 헤지액면
  - 현금흐름 : 종목별 이표 스케줄 + 한국 관행식 단가계산 + 선도수익률 뉴턴해
  - 이론가   : 표준물 가격, 이론 선물가, 선물 내재수익률, 베이시스
  - 손익     : 만기 손익(바스켓 1단위 = 선물 1계약)
  - 시나리오 : 만기 수익률 평행이동 -100 ~ +100bp 손익표
  - 연동가이드 / README

가격산식: 한국 채권시장 관행 (단수기간 단리 + 이후 반기복리)
    단가(더티) = [ Σ CF_k / (1+y/2)^(k-1) ] / (1 + y/2 × d/T)
표준물     : 잔존 정확히 n년 → 단수기간 없음 → 폐쇄형
    P = (c×100/y)(1-v^N) + 100·v^N,  v = 1/(1+y/2), N = 2n
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
import datetime as dt

OUT = "/home/user/vmo/ktb_arb/국채선물_차익거래.xlsx"

# ──────────────────────────── 스타일 ────────────────────────────
FONT = "Arial"
F_TITLE = Font(name=FONT, size=14, bold=True, color="FFFFFF")
F_SEC = Font(name=FONT, size=11, bold=True, color="FFFFFF")
F_HDR = Font(name=FONT, size=10, bold=True)
F_LBL = Font(name=FONT, size=10)
F_IN = Font(name=FONT, size=10, color="0000FF")          # 하드코딩 입력
F_LIVE = Font(name=FONT, size=10, bold=True, color="C00000")  # 실시간(RTD) 셀
F_CALC = Font(name=FONT, size=10)                        # 계산식
F_LINK = Font(name=FONT, size=10, color="008000")        # 타시트 참조
F_NOTE = Font(name=FONT, size=9, italic=True, color="808080")
F_KEY = Font(name=FONT, size=11, bold=True)

FILL_TITLE = PatternFill("solid", fgColor="1F3864")
FILL_SEC = PatternFill("solid", fgColor="4472C4")
FILL_IN = PatternFill("solid", fgColor="FFFF99")         # 사용자 입력
FILL_LIVE = PatternFill("solid", fgColor="DDEBF7")       # 실시간 연결
FILL_KEY = PatternFill("solid", fgColor="E2EFDA")        # 핵심 결과
FILL_HDR = PatternFill("solid", fgColor="D9D9D9")

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

N_WON = '#,##0'
N_WON2 = '#,##0.0'
N_PX = '0.0000'
N_PX2 = '0.00'
N_YLD = '0.0000'
N_BP = '0.00'
N_DATE = 'yyyy-mm-dd'
N_INT = '#,##0'


def as_text(cell):
    """'='로 시작하는 예시/설명 문자열을 수식이 아닌 텍스트로 저장한다."""
    if isinstance(cell.value, str) and cell.value.startswith("="):
        cell.data_type = "s"
    return cell


def title(ws, text, width=8):
    ws["A1"] = text
    ws["A1"].font = F_TITLE
    ws["A1"].fill = FILL_TITLE
    for c in range(1, width + 1):
        ws.cell(row=1, column=c).fill = FILL_TITLE
    ws.row_dimensions[1].height = 22


def section(ws, row, text, width=8):
    ws.cell(row=row, column=1, value=text).font = F_SEC
    for c in range(1, width + 1):
        ws.cell(row=row, column=c).fill = FILL_SEC


def put(ws, row, label, value, unit="", fmt=None, style="calc", note="",
        lab_col=1, val_col=2, unit_col=3, note_col=4):
    ws.cell(row=row, column=lab_col, value=label).font = F_LBL
    c = ws.cell(row=row, column=val_col, value=value)
    if style == "in":
        c.font, c.fill = F_IN, FILL_IN
    elif style == "live":
        c.font, c.fill = F_LIVE, FILL_LIVE
    elif style == "link":
        c.font = F_LINK
    elif style == "key":
        c.font, c.fill = F_KEY, FILL_KEY
    else:
        c.font = F_CALC
    if fmt:
        c.number_format = fmt
    c.border = BOX
    if unit:
        ws.cell(row=row, column=unit_col, value=unit).font = F_NOTE
    if note:
        ws.cell(row=row, column=note_col, value=note).font = F_NOTE
    return c


wb = Workbook()

# ════════════════════════════════════════════════════════════════
# 0. README
# ════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "README"
title(ws, "국채선물 차익거래 실시간 모니터  ―  사용설명서", 6)
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 118

readme = [
    ("s", "1. 이 파일이 계산하는 것"),
    ("", "한국 국채선물은 실물인수도가 아닌 '현금결제' 상품이다. 최종결제기준가격은 최종거래일에 국고채전문딜러들이"),
    ("", "제출한 바스켓(최종결제기준채권) 종목별 수익률을 집계·평균하여, 표면금리 5% / 6개월 이표 / 잔존 n년의"),
    ("", "'표준물(가상채권)' 할인율로 적용해 산출한다. 따라서 차익거래 손익은 다음과 같이 분해된다."),
    ("", ""),
    ("", "    만기손익 = [선물 leg] (진입가 − 표준물가격(바스켓 평균수익률)) × 1포인트가치 × 계약수 × 방향"),
    ("", "             + [현물 leg] Σ (만기 처분단가 − 선도 목표단가) / 100 × 보유액면 × 방향"),
    ("", "             − 거래비용"),
    ("", ""),
    ("", "선도 목표단가 = 현물단가 + 조달이자 − 기간중 쿠폰 재투자원리금  (= 캐리 반영 선도가격)"),
    ("", "각 종목의 선도 BPV 기준으로 헤지액면을 잡으면 [현물 leg]가 수익률 변동을 상쇄하여,"),
    ("", "만기손익 ≈ 순베이시스(= 시장선물가 − 이론선물가) × 1포인트가치 − 비용 으로 '락인'된다."),
    ("s", "2. 사용 순서"),
    ("", "① [실시간] 시트의 붉은 글씨 셀에 사용 중인 데이터 벤더의 RTD/DDE 수식을 붙여넣는다. (→ 연동가이드 시트)"),
    ("", "② [설정] 시트에서 대상 선물(3/5/10/30년), 결제월, 조달금리, 계약수, 거래비용을 입력한다. (노란 셀)"),
    ("", "③ [바스켓] 시트 4~10행에 해당 결제월의 최종결제기준채권 3종목(종목코드/발행일/만기일/표면금리/가중치)을 입력한다."),
    ("", "④ [이론가] 시트에서 베이시스·순차익을, [손익] 시트에서 만기손익을, [시나리오] 시트에서 헤지 유효성을 확인한다."),
    ("s", "3. 셀 색상 규칙"),
    ("", "노란 배경 + 파란 글씨 = 사용자가 직접 입력하는 값        하늘색 배경 + 붉은 글씨 = 실시간(RTD) 연결 셀"),
    ("", "초록 글씨 = 다른 시트 참조                                연두색 배경 = 핵심 결과값"),
    ("s", "4. 계산 규약 (변경 시 [현금흐름] 시트 수식 수정 필요)"),
    ("", "· 채권단가: 한국 관행식 = [Σ CF_k /(1+y/2)^(k-1)] / (1 + y/2 × d/T).  단수기간 단리, 이후 반기복리."),
    ("", "  이 식의 결과는 경과이자를 포함한 '더티(단가)' 이며, 본 파일의 모든 손익계산은 더티 기준으로 일관한다."),
    ("", "  (경과이자·클린가는 참고용으로만 표시하며 실제일수/365 단리로 산출한다.)"),
    ("", "· 표준물: 잔존 정확히 n년 → 단수기간 없음 → P = (5/y)(1−v^N) + 100·v^N,  v=1/(1+y/2), N=2n."),
    ("", "· 선도수익률: 선도 목표단가에 대해 뉴턴법 2회 반복으로 역산 (수렴 잔차를 [바스켓] 38행에서 확인)."),
    ("", "· 시나리오/만기 손익의 현물 평가는 이표 스케줄 전체를 재할인하는 정밀 재평가이며 근사를 쓰지 않는다."),
    ("", "  ([바스켓] 39·41행의 BPV/2차미분은 헤지비율 산출과 참고용 지표로만 쓰인다.)"),
    ("s", "5. 이 파일이 '반영하지 않는' 것 (수동 판단 필요)"),
    ("", "· 휴일 캘린더: 최종결제일/현물결제일은 WORKDAY() 기반이라 공휴일 미반영. [설정]에서 수동 입력으로 덮어쓸 것."),
    ("", "· 선물 일일정산(변동증거금)의 재투자·조달 이자효과(테일 조정). 잔존기간이 길고 변동성이 크면 별도 고려."),
    ("", "· 최종결제기준가격은 딜러 제출 수익률의 최고·최저 제외 평균이므로, 만기일에 본인이 실제 체결하는"),
    ("", "  현물 매도수익률과 괴리가 발생할 수 있다(가장 큰 잔존 리스크). 슬리피지 가정을 [설정] 거래비용에 반영할 것."),
    ("", "· 레포 롤오버 리스크, 담보 haircut, 기관 유형별 이자소득 원천징수, 회계상 평가손익 구분."),
    ("s", "6. 출처 / 사용자 제공값"),
    ("", "· 계약사양(거래단위·호가단위·최종거래일 규칙), 결제월별 최종결제기준채권 구성: 한국거래소(KRX) 공시 기준."),
    ("", "  → 사용 전 반드시 KRX 파생상품 상품명세 및 해당 결제월 바스켓 공지로 재확인할 것."),
    ("", "· 초기 입력값(바스켓 종목, 수익률, 선물가격, 조달금리)은 형식 예시용 더미 데이터이며 실거래 값이 아니다."),
]
r = 3
for kind, txt in readme:
    if kind == "s":
        ws.cell(row=r, column=2, value=txt).font = Font(name=FONT, size=11, bold=True, color="1F3864")
        r += 1
    else:
        ws.cell(row=r, column=2, value=txt).font = F_LBL
        r += 1

# ════════════════════════════════════════════════════════════════
# 1. 실시간
# ════════════════════════════════════════════════════════════════
lv = wb.create_sheet("실시간")
title(lv, "실시간 데이터 연결 시트  ―  붉은 셀에 벤더 RTD/DDE 수식을 붙여넣으세요", 6)
for col, w in zip("ABCDEF", [30, 22, 14, 46, 34, 4]):
    lv.column_dimensions[col].width = w

for c, h in zip("ABCDE", ["항목", "종목코드/심볼", "값", "연동 수식 예시 (연합인포맥스)", "비고"]):
    cell = lv[f"{c}3"]
    cell.value = h
    cell.font = F_HDR
    cell.fill = FILL_HDR
    cell.border = BOX

LIVE_ROWS = [
    (5,  "국채선물 현재가",            "KTB10 최근월물", 117.52, "=IMFX(\"F 10YKTB\",\"현재가\")",   "가격 100 기준"),
    (6,  "국채선물 매수1호가(Bid)",    "KTB10 최근월물", 117.51, "=IMFX(\"F 10YKTB\",\"매수1호가\")", "선물 매도 시 체결가"),
    (7,  "국채선물 매도1호가(Ask)",    "KTB10 최근월물", 117.53, "=IMFX(\"F 10YKTB\",\"매도1호가\")", "선물 매수 시 체결가"),
    (8,  "국채선물 미드",              "",               None,   "=AVERAGE(C6:C7)",                   "자동계산"),
    (10, "바스켓1 Bid수익률(딜러매수)", "KR103502GE97",  2.9600, "=IMFX(\"KR103502GE97\",\"매수수익률\")", "내가 매도할 때 적용 (%)"),
    (11, "바스켓1 Ask수익률(딜러매도)", "KR103502GE97",  2.9500, "=IMFX(\"KR103502GE97\",\"매도수익률\")", "내가 매수할 때 적용 (%)"),
    (12, "바스켓2 Bid수익률(딜러매수)", "KR103503GF95",  2.9700, "=IMFX(\"KR103503GF95\",\"매수수익률\")", "(%)"),
    (13, "바스켓2 Ask수익률(딜러매도)", "KR103503GF95",  2.9600, "=IMFX(\"KR103503GF95\",\"매도수익률\")", "(%)"),
    (14, "바스켓3 Bid수익률(딜러매수)", "KR103504GH93",  2.9800, "=IMFX(\"KR103504GH93\",\"매수수익률\")", "(%)"),
    (15, "바스켓3 Ask수익률(딜러매도)", "KR103504GH93",  2.9700, "=IMFX(\"KR103504GH93\",\"매도수익률\")", "(%)"),
    (17, "RP(레포) 조달금리",          "",               2.6000, "=IMFX(\"RP1D\",\"금리\")",          "[설정]에서 '실시간' 선택 시 사용 (%)"),
    (18, "CD 91일",                    "",               2.7500, "=IMFX(\"CD91\",\"금리\")",          "참고 (%)"),
    (19, "국고 3년 지표",              "",               2.7000, "=IMFX(\"KTB3Y\",\"수익률\")",       "참고 (%)"),
]
for row, name, sym, val, ex, note in LIVE_ROWS:
    lv.cell(row=row, column=1, value=name).font = F_LBL
    lv.cell(row=row, column=2, value=sym).font = F_IN
    lv.cell(row=row, column=2).fill = FILL_IN
    c = lv.cell(row=row, column=3)
    if row == 8:
        c.value = "=AVERAGE(C6:C7)"
        c.font = F_CALC
    else:
        c.value = val
        c.font, c.fill = F_LIVE, FILL_LIVE
    c.number_format = N_PX if row <= 8 else N_YLD
    c.border = BOX
    as_text(lv.cell(row=row, column=4, value=ex)).font = F_NOTE
    lv.cell(row=row, column=5, value=note).font = F_NOTE

lv["A21"] = "갱신시각"
lv["A21"].font = F_LBL
lv["C21"] = "=NOW()"
lv["C21"].number_format = "yyyy-mm-dd hh:mm:ss"
lv["C21"].font = F_CALC
lv["A23"] = ("※ 현재 들어있는 숫자는 형식 확인용 더미값입니다. 실거래 전 반드시 벤더 수식으로 교체하세요. "
             "수식 예시는 [연동가이드] 시트 참조.")
lv["A23"].font = F_NOTE
lv["A24"] = ("※ 채권 수익률은 '딜러 매수호가(Bid)=높은 수익률/낮은 가격', '딜러 매도호가(Ask)=낮은 수익률/높은 가격' 입니다. "
             "내가 바스켓을 매수하면 Ask수익률이 체결 기준입니다.")
lv["A24"].font = F_NOTE
lv["C5"].comment = Comment("실시간 연결 셀. 벤더 RTD/DDE 수식으로 교체하십시오.\n초기값은 더미 데이터입니다.", "설계")

# ════════════════════════════════════════════════════════════════
# 2. 설정
# ════════════════════════════════════════════════════════════════
cf = wb.create_sheet("설정")
title(cf, "설정  ―  노란 셀만 입력하십시오", 6)
for col, w in zip("ABCDEF", [34, 18, 12, 52, 10, 10]):
    cf.column_dimensions[col].width = w

section(cf, 3, "[1] 선물 계약 사양", 6)
put(cf, 4, "대상 선물", "10년", "", None, "in", "3년 / 5년 / 10년 / 30년")
# 보조 매핑표
cf["F4"], cf["G4"] = "3년", 3
cf["F5"], cf["G5"] = "5년", 5
cf["F6"], cf["G6"] = "10년", 10
cf["F7"], cf["G7"] = "30년", 30
for rr in range(4, 8):
    cf.cell(row=rr, column=6).font = F_NOTE
    cf.cell(row=rr, column=7).font = F_NOTE
cf["F3"] = "▼ 매핑표(수정금지)"
cf["F3"].font = F_NOTE
put(cf, 5, "표준물 만기", "=IFERROR(VLOOKUP($B$4,$F$4:$G$7,2,FALSE),10)", "년", N_INT, "calc")
put(cf, 6, "표준물 표면금리", 0.05, "", "0.00%", "in", "KRX 표준물 규격: 연 5%")
put(cf, 7, "이표주기", 2, "회/년", N_INT, "in", "6개월 이표")
put(cf, 8, "거래단위(액면)", 100000000, "원", N_WON, "in", "KRX 명세 확인 필수")
put(cf, 9, "가격 1.00포인트 가치", "=B8/100", "원", N_WON, "calc")
put(cf, 10, "호가단위(틱)", 0.01, "", "0.00", "in")
put(cf, 11, "1틱 가치", "=B10*B9", "원", N_WON, "calc")

section(cf, 13, "[2] 결제 일정", 6)
put(cf, 14, "결제 연도", 2026, "", N_INT, "in")
put(cf, 15, "결제 월", 9, "", N_INT, "in", "3 / 6 / 9 / 12")
put(cf, 16, "최종거래일(자동: 세번째 화요일)",
    "=DATE(B14,B15,1)+MOD(2-WEEKDAY(DATE(B14,B15,1),2)+7,7)+14", "", N_DATE, "calc")
put(cf, 17, "최종거래일(수동 override)", None, "", N_DATE, "in", "공휴일 등으로 조정 시 입력")
put(cf, 18, "▶ 적용 최종거래일", '=IF(B17="",B16,B17)', "", N_DATE, "key")
put(cf, 19, "최종결제일(자동: 익영업일)", "=WORKDAY(B18,1)", "", N_DATE, "calc")
put(cf, 20, "최종결제일(수동 override)", None, "", N_DATE, "in")
put(cf, 21, "▶ 적용 최종결제일 (H)", '=IF(B20="",B19,B20)', "", N_DATE, "key", "현물 청산 결제일 = 손익 확정 시점")

section(cf, 23, "[3] 평가일 / 금리", 6)
put(cf, 24, "평가일(오늘)", "=TODAY()", "", N_DATE, "calc")
put(cf, 25, "현물 결제일(자동: T+1)", "=WORKDAY(B24,1)", "", N_DATE, "calc", "국고채 장외 익일결제 관행")
put(cf, 26, "현물 결제일(수동 override)", None, "", N_DATE, "in")
put(cf, 27, "▶ 적용 현물결제일 (S0)", '=IF(B26="",B25,B26)', "", N_DATE, "key")
put(cf, 28, "보유일수 (S0 → H)", "=B21-B27", "일", N_INT, "key")
put(cf, 29, "조달금리 소스", "수동", "", None, "in", "수동 / 실시간")
put(cf, 30, "조달금리(수동)", 2.60, "%", N_YLD, "in", "레포(RP) 실효 조달금리")
put(cf, 31, "▶ 적용 조달금리", '=IF(B29="실시간",실시간!C17,B30)', "%", N_YLD, "key")
put(cf, 32, "적용 조달금리(소수)", "=B31/100", "", "0.000000", "calc")
put(cf, 33, "쿠폰 재투자금리", "=B31", "%", N_YLD, "calc", "기본: 조달금리와 동일")
put(cf, 34, "적용 재투자금리(소수)", "=B33/100", "", "0.000000", "calc")
put(cf, 35, "이자 일수기준", 365, "일", N_INT, "in", "actual/365 단리")

section(cf, 37, "[4] 포지션 / 헤지", 6)
put(cf, 38, "포지션 방향", "선물매도+현물매수", "", None, "in", "선물매도+현물매수 / 선물매수+현물매도")
put(cf, 39, "방향계수", '=IF(B38="선물매도+현물매수",1,-1)', "", N_INT, "calc")
put(cf, 40, "계약수 (Q)", 10, "계약", N_INT, "in")
put(cf, 41, "선물 진입가격", None, "", N_PX2, "in", "미입력 시 현재가 사용")
put(cf, 42, "▶ 적용 선물 진입가", '=IF(B41="",실시간!C5,B41)', "", N_PX2, "key")
put(cf, 43, "진입일", None, "", N_DATE, "in", "미입력 시 오늘")
put(cf, 44, "▶ 적용 진입일", '=IF(B43="",B24,B43)', "", N_DATE, "key")
put(cf, 45, "헤지액면 자동사용", "예", "", None, "in", "예 = 선도BPV 기준 이론액면 / 아니오 = 실제 보유액면")
put(cf, 46, "수익률 적용 기준", "호가", "", None, "in", "호가 = 체결가능측 / 미드 = 중간값")

section(cf, 48, "[5] 거래비용 (계약당)", 6)
put(cf, 49, "선물 왕복 수수료", 3000, "원/계약", N_WON, "in", "증권사 위탁+거래소 수수료")
put(cf, 50, "채권 매매비용(왕복)", 1.0, "bp", N_BP, "in", "장외 스프레드·슬리피지 가정. 호가수익률 사용 시 중복 주의")
put(cf, 51, "채권 매매비용", "=B50*0.0001*B8", "원/계약", N_WON, "calc")
put(cf, 52, "기타 비용(결제·최종결제 슬리피지 등)", 0, "원/계약", N_WON, "in")
put(cf, 53, "▶ 총 거래비용", "=B49+B51+B52", "원/계약", N_WON, "key")

for dv_range, formula in [
    ("B4", '"3년,5년,10년,30년"'),
    ("B29", '"수동,실시간"'),
    ("B38", '"선물매도+현물매수,선물매수+현물매도"'),
    ("B45", '"예,아니오"'),
    ("B46", '"호가,미드"'),
]:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    cf.add_data_validation(dv)
    dv.add(cf[dv_range])

cf["A55"] = ("※ B8(거래단위), B10(호가단위), B16(최종거래일 규칙)은 KRX 파생상품 상품명세를 기준으로 한 값입니다. "
             "특히 30년물은 명세를 반드시 재확인하십시오.")
cf["A55"].font = F_NOTE
cf["A56"] = "※ B50(채권 매매비용)과 [실시간]의 Bid/Ask 호가를 동시에 쓰면 스프레드가 이중 계상됩니다. B46='호가'면 B50=0 권장."
cf["A56"].font = F_NOTE

# ════════════════════════════════════════════════════════════════
# 3. 현금흐름 (계산 엔진)
# ════════════════════════════════════════════════════════════════
cs = wb.create_sheet("현금흐름")
title(cs, "현금흐름 스케줄 & 단가 계산 엔진  (직접 수정 금지)", 10)

NPER = 62                 # 31년치 반기 이표 (30년물 커버)
SCH0, SCH1 = 15, 15 + NPER - 1     # 15 ~ 76
BLOCK_STARTS = [1, 43, 85]         # A, AQ, CG  (블록폭 40)

SCH_HDR = ["k", "이표일", "CF(100기준)", "현물포함", "현물idx",
           "PV@y", "PV@y+1bp", "PV@y-1bp", "선도포함", "선도idx",
           "PV_H@y0", "PV_H@y0+", "PV_H@y0-", "PV_H@y1", "PV_H@y2",
           "PV_H@y2+", "PV_H@y2-", "쿠폰재투자FV",
           "PV_H@손익shift"] + [f"PV_H@시나리오{j+1}" for j in range(21)]

SCALAR1 = [
    (2,  "만기일",            "=바스켓!{bc}7",   N_DATE),
    (3,  "발행일",            "=바스켓!{bc}6",   N_DATE),
    (4,  "표면금리(소수)",     "=바스켓!{bc}8/100", "0.00000"),
    (5,  "반기 쿠폰액",        "={v}4*100/바스켓!{bc}9", N_PX),
    (6,  "현물결제일 S0",      "=설정!$B$27",     N_DATE),
    (7,  "최종결제일 H",       "=설정!$B$21",     N_DATE),
    (8,  "현물 수익률(소수)",  "=바스켓!{bc}17",  "0.000000"),
    (9,  "조달금리(소수)",     "=설정!$B$32",     "0.000000"),
    (10, "재투자금리(소수)",   "=설정!$B$34",     "0.000000"),
    (11, "이자일수기준",       "=설정!$B$35",     N_INT),
]

for bi, s in enumerate(BLOCK_STARTS):
    C = lambda off: get_column_letter(s + off)
    v = C(1)      # 스칼라 값 열 (= 이표일 열과 동일 열, 다른 행)
    bc = "CDE"[bi]

    cs.cell(row=12, column=s, value=f"■ 종목{bi+1} 이표 스케줄").font = Font(
        name=FONT, size=11, bold=True, color="1F3864")
    cs.column_dimensions[C(0)].width = 22
    cs.column_dimensions[C(1)].width = 15
    for off in range(2, 40):
        cs.column_dimensions[C(off)].width = 12

    for row, lab, fml, fmt in SCALAR1:
        cs.cell(row=row, column=s, value=lab).font = F_LBL
        c = cs.cell(row=row, column=s + 1, value=fml.format(bc=bc, v=v))
        c.font, c.number_format, c.border = F_LINK, fmt, BOX

    # 스케줄 헤더
    for off, h in enumerate(SCH_HDR):
        c = cs.cell(row=14, column=s + off, value=h)
        c.font, c.fill, c.border = F_HDR, FILL_HDR, BOX
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for i in range(NPER):
        R = SCH0 + i
        cols = {n: C(n) for n in range(18)}
        vals = {
            0:  i + 1,
            1:  f"=EDATE(${v}$2,-6*({NPER}-${cols[0]}{R}))",
            2:  f'=IF(AND({cols[1]}{R}>${v}$3,{cols[1]}{R}<=${v}$2),IF({cols[1]}{R}=${v}$2,${v}$5+100,${v}$5),0)',
            3:  f'=IF(AND(${cols[2]}{R}>0,{cols[1]}{R}>${v}$6),1,0)',
            4:  f'=IF({cols[3]}{R}=1,SUM(${cols[3]}${SCH0}:{cols[3]}{R}),"")',
            5:  f'=IF({cols[3]}{R}=1,${cols[2]}{R}/(1+${v}$8/2)^({cols[4]}{R}-1),0)',
            6:  f'=IF({cols[3]}{R}=1,${cols[2]}{R}/(1+(${v}$8+0.0001)/2)^({cols[4]}{R}-1),0)',
            7:  f'=IF({cols[3]}{R}=1,${cols[2]}{R}/(1+(${v}$8-0.0001)/2)^({cols[4]}{R}-1),0)',
            8:  f'=IF(AND(${cols[2]}{R}>0,{cols[1]}{R}>${v}$7),1,0)',
            9:  f'=IF({cols[8]}{R}=1,SUM(${cols[8]}${SCH0}:{cols[8]}{R}),"")',
            10: f'=IF({cols[8]}{R}=1,${cols[2]}{R}/(1+${v}$104/2)^({cols[9]}{R}-1),0)',
            11: f'=IF({cols[8]}{R}=1,${cols[2]}{R}/(1+(${v}$104+0.0001)/2)^({cols[9]}{R}-1),0)',
            12: f'=IF({cols[8]}{R}=1,${cols[2]}{R}/(1+(${v}$104-0.0001)/2)^({cols[9]}{R}-1),0)',
            13: f'=IF({cols[8]}{R}=1,${cols[2]}{R}/(1+${v}$109/2)^({cols[9]}{R}-1),0)',
            14: f'=IF({cols[8]}{R}=1,${cols[2]}{R}/(1+${v}$111/2)^({cols[9]}{R}-1),0)',
            15: f'=IF({cols[8]}{R}=1,${cols[2]}{R}/(1+(${v}$111+0.0001)/2)^({cols[9]}{R}-1),0)',
            16: f'=IF({cols[8]}{R}=1,${cols[2]}{R}/(1+(${v}$111-0.0001)/2)^({cols[9]}{R}-1),0)',
            17: f'=IF(AND(${cols[2]}{R}>0,{cols[1]}{R}>${v}$6,{cols[1]}{R}<=${v}$7),'
                f'${cols[2]}{R}*(1+${v}$10*(${v}$7-{cols[1]}{R})/365),0)',
        }
        # 정밀 재평가용 PV 열: 손익 단일 shift(off=18) + 시나리오 21개(off=19..39)
        vals[18] = (f'=IF({cols[8]}{R}=1,${cols[2]}{R}/'
                    f'(1+(${v}$111+손익!$B$24/10000)/2)^({cols[9]}{R}-1),0)')
        for j in range(21):
            vals[19 + j] = (f'=IF({cols[8]}{R}=1,${cols[2]}{R}/'
                            f'(1+(${v}$111+시나리오!$A${4+j}/10000)/2)^({cols[9]}{R}-1),0)')
        for off, val in vals.items():
            c = cs.cell(row=R, column=s + off, value=val)
            c.font = F_CALC
            if off == 1:
                c.number_format = N_DATE
            elif off in (0, 3, 4, 8, 9):
                c.number_format = N_INT
            else:
                c.number_format = "0.000000"

    sc = lambda off: f"{C(off)}{SCH0}:{C(off)}{SCH1}"
    SCALAR2 = [
        (78,  "── 현물 평가 (S0) ──", None, None),
        (79,  "잔존 이표 횟수 n",        f"=SUM({sc(3)})", N_INT),
        (80,  "다음 이표일",             f"=INDEX({sc(1)},MATCH(1,{sc(3)},0))", N_DATE),
        (81,  "직전 이표일",             f"=EDATE({v}80,-6)", N_DATE),
        (82,  "d (S0→다음이표일)",       f"={v}80-{v}6", N_INT),
        (83,  "T (이표기간 일수)",        f"={v}80-{v}81", N_INT),
        (84,  "단수기간 할인계수",        f"=1+{v}8/2*{v}82/{v}83", "0.000000"),
        (85,  "단가(더티) P0",           f"=SUM({sc(5)})/{v}84", "0.000000"),
        (86,  "P0 @ y+1bp",             f"=SUM({sc(6)})/(1+({v}8+0.0001)/2*{v}82/{v}83)", "0.000000"),
        (87,  "P0 @ y-1bp",             f"=SUM({sc(7)})/(1+({v}8-0.0001)/2*{v}82/{v}83)", "0.000000"),
        (88,  "BPV (포인트/1bp)",        f"=({v}87-{v}86)/2", "0.000000"),
        (89,  "수정듀레이션(년)",         f"={v}88*10000/{v}85", "0.0000"),
        (90,  "경과일수(참고)",           f"={v}6-{v}81", N_INT),
        (91,  "경과이자(참고)",           f"={v}4*100*{v}90/365", "0.000000"),
        (92,  "클린가(참고)",             f"={v}85-{v}91", "0.000000"),
        (94,  "── 선도 평가 (H) ──", None, None),
        (95,  "선도 잔존 이표횟수",       f"=SUM({sc(8)})", N_INT),
        (96,  "H 이후 다음 이표일",       f"=INDEX({sc(1)},MATCH(1,{sc(8)},0))", N_DATE),
        (97,  "그 직전 이표일",           f"=EDATE({v}96,-6)", N_DATE),
        (98,  "d_H",                    f"={v}96-{v}7", N_INT),
        (99,  "T_H",                    f"={v}96-{v}97", N_INT),
        (100, "기간중 쿠폰 재투자FV",     f"=SUM({sc(17)})", "0.000000"),
        (101, "조달일수 (S0→H)",         f"={v}7-{v}6", N_INT),
        (102, "조달이자",                f"={v}85*{v}9*{v}101/365", "0.000000"),
        (103, "선도 목표단가(더티)",      f"={v}85+{v}102-{v}100", "0.000000"),
        (104, "y0 (초기추정)",           f"={v}8", "0.00000000"),
        (105, "P_H(y0)",                f"=SUM({sc(10)})/(1+{v}104/2*{v}98/{v}99)", "0.000000"),
        (106, "P_H(y0+1bp)",            f"=SUM({sc(11)})/(1+({v}104+0.0001)/2*{v}98/{v}99)", "0.000000"),
        (107, "P_H(y0-1bp)",            f"=SUM({sc(12)})/(1+({v}104-0.0001)/2*{v}98/{v}99)", "0.000000"),
        (108, "|dP/dy| (기울기)",        f"=({v}107-{v}106)/0.0002", "0.000000"),
        (109, "y1 (뉴턴 1회)",           f"={v}104+({v}105-{v}103)/{v}108", "0.00000000"),
        (110, "P_H(y1)",                f"=SUM({sc(13)})/(1+{v}109/2*{v}98/{v}99)", "0.000000"),
        (111, "y2 (뉴턴 2회) ▶ 선도수익률", f"={v}109+({v}110-{v}103)/{v}108", "0.00000000"),
        (112, "P_H(y2)",                f"=SUM({sc(14)})/(1+{v}111/2*{v}98/{v}99)", "0.000000"),
        (113, "수렴 잔차",               f"={v}112-{v}103", "0.00E+00"),
        (114, "P_H(y2+1bp)",            f"=SUM({sc(15)})/(1+({v}111+0.0001)/2*{v}98/{v}99)", "0.000000"),
        (115, "P_H(y2-1bp)",            f"=SUM({sc(16)})/(1+({v}111-0.0001)/2*{v}98/{v}99)", "0.000000"),
        (116, "선도 BPV (포인트/1bp)",    f"=({v}115-{v}114)/2", "0.000000"),
        (117, "선도 2차미분 P''",        f"=({v}114-2*{v}112+{v}115)/(0.0001^2)", "0.00"),
        (118, "선도 잔존만기(년)",        f"=({v}2-{v}7)/365", "0.0000"),
        (120, "── 정밀 재평가 (근사 없음) ──", None, None),
        (121, "P_H @ y2+손익shift",
         f"=SUM({sc(18)})/(1+({v}111+손익!$B$24/10000)/2*{v}98/{v}99)", "0.000000"),
    ] + [
        (123 + j, f"P_H @ y2+시나리오{j+1}",
         f"=SUM({sc(19+j)})/(1+({v}111+시나리오!$A${4+j}/10000)/2*{v}98/{v}99)", "0.000000")
        for j in range(21)
    ]
    for row, lab, fml, fmt in SCALAR2:
        c0 = cs.cell(row=row, column=s, value=lab)
        if fml is None:
            c0.font = Font(name=FONT, size=10, bold=True, color="1F3864")
            continue
        c0.font = F_LBL
        c = cs.cell(row=row, column=s + 1, value=fml)
        c.font, c.number_format, c.border = F_CALC, fmt, BOX

cs.freeze_panes = "A15"

# ════════════════════════════════════════════════════════════════
# 4. 바스켓
# ════════════════════════════════════════════════════════════════
bs = wb.create_sheet("바스켓")
title(bs, "바스켓 채권 (최종결제기준채권)  ―  노란 셀 입력", 7)
for col, w in zip("ABCDEFG", [32, 14, 20, 20, 20, 22, 44]):
    bs.column_dimensions[col].width = w

for col, h in zip("ABCDEFG", ["항목", "단위", "종목1", "종목2", "종목3", "합계/검증", "비고"]):
    c = bs[f"{col}3"]
    c.value, c.font, c.fill, c.border = h, F_HDR, FILL_HDR, BOX

DUMMY = [
    ("KR103502GE97", "국고03000-3506(25-3)", dt.date(2025, 6, 10), dt.date(2035, 6, 10), 3.000),
    ("KR103503GF95", "국고02875-3509(25-6)", dt.date(2025, 9, 10), dt.date(2035, 9, 10), 2.875),
    ("KR103504GH93", "국고03125-3512(25-9)", dt.date(2025, 12, 10), dt.date(2035, 12, 10), 3.125),
]
CFV = {"C": "B", "D": "AR", "E": "CH"}   # 바스켓 열 → 현금흐름 스칼라 값 열


def brow(row, label, unit, per_bond, fmt=None, style="calc", total=None, note=""):
    bs.cell(row=row, column=1, value=label).font = F_LBL
    bs.cell(row=row, column=2, value=unit).font = F_NOTE
    for i, col in enumerate("CDE"):
        val = per_bond(col, i) if callable(per_bond) else per_bond
        c = bs[f"{col}{row}"]
        c.value = val
        if style == "in":
            c.font, c.fill = F_IN, FILL_IN
        elif style == "live":
            c.font = F_LINK
        elif style == "key":
            c.font, c.fill = F_KEY, FILL_KEY
        else:
            c.font = F_CALC
        if fmt:
            c.number_format = fmt
        c.border = BOX
    if total:
        c = bs[f"F{row}"]
        c.value, c.font, c.border = total, F_KEY, BOX
        if fmt:
            c.number_format = fmt
    if note:
        bs.cell(row=row, column=7, value=note).font = F_NOTE


brow(4, "종목코드", "", lambda col, i: DUMMY[i][0], None, "in")
brow(5, "종목명", "", lambda col, i: DUMMY[i][1], None, "in")
brow(6, "발행일", "", lambda col, i: DUMMY[i][2], N_DATE, "in", note="이표일 생성 기준")
brow(7, "만기일", "", lambda col, i: DUMMY[i][3], N_DATE, "in", note="이표일을 만기에서 6개월씩 역산")
brow(8, "표면금리", "%", lambda col, i: DUMMY[i][4], N_YLD, "in")
brow(9, "이표주기", "회/년", 2, N_INT, "in")
brow(10, "가중치 w", "", "=1/3", "0.0000", "in", total="=SUM(C10:E10)",
     note="KRX 최종결제기준가격은 종목별 수익률의 단순평균 → 기본 1/3. 미사용 종목은 0")

section(bs, 12, "[실시간 수익률]", 7)
brow(13, "Bid 수익률 (딜러 매수호가)", "%", lambda col, i: f"=실시간!C{10+2*i}", N_YLD, "live")
brow(14, "Ask 수익률 (딜러 매도호가)", "%", lambda col, i: f"=실시간!C{11+2*i}", N_YLD, "live")
brow(15, "Mid 수익률", "%", lambda col, i: f"=AVERAGE({col}13:{col}14)", N_YLD)
brow(16, "▶ 적용 수익률", "%", lambda col, i:
     f'=IF(설정!$B$46="미드",{col}15,IF(설정!$B$39=1,{col}14,{col}13))', N_YLD, "key",
     note="현물 매수(방향=1)면 Ask, 매도면 Bid 적용")
brow(17, "적용 수익률(소수)", "", lambda col, i: f"={col}16/100", "0.000000")

section(bs, 19, "[현물 평가 — 결제일 S0]", 7)
brow(20, "잔존 이표 횟수", "회", lambda col, i: f"=현금흐름!${CFV[col]}$79", N_INT)
brow(21, "다음 이표일", "", lambda col, i: f"=현금흐름!${CFV[col]}$80", N_DATE)
brow(22, "직전 이표일", "", lambda col, i: f"=현금흐름!${CFV[col]}$81", N_DATE)
brow(23, "d (결제일→다음이표일)", "일", lambda col, i: f"=현금흐름!${CFV[col]}$82", N_INT)
brow(24, "T (이표기간 일수)", "일", lambda col, i: f"=현금흐름!${CFV[col]}$83", N_INT)
brow(25, "▶ 단가 (더티, 100 기준)", "", lambda col, i: f"=현금흐름!${CFV[col]}$85", "0.0000", "key")
brow(26, "경과이자 (참고)", "", lambda col, i: f"=현금흐름!${CFV[col]}$91", "0.0000")
brow(27, "클린가 (참고)", "", lambda col, i: f"=현금흐름!${CFV[col]}$92", "0.0000")
brow(28, "현물 BPV", "포인트/1bp", lambda col, i: f"=현금흐름!${CFV[col]}$88", "0.000000")
brow(29, "현물 BPV (액면 1억 기준)", "원/1bp", lambda col, i: f"={col}28*설정!$B$9", N_WON2)
brow(30, "수정듀레이션", "년", lambda col, i: f"=현금흐름!${CFV[col]}$89", "0.0000")

section(bs, 32, "[선도 평가 — 최종결제일 H]", 7)
brow(33, "보유일수 (S0→H)", "일", "=설정!$B$28", N_INT)
brow(34, "기간중 쿠폰 재투자FV", "", lambda col, i: f"=현금흐름!${CFV[col]}$100", "0.0000")
brow(35, "조달이자", "", lambda col, i: f"=현금흐름!${CFV[col]}$102", "0.0000")
brow(36, "▶ 선도 목표단가 (더티)", "", lambda col, i: f"=현금흐름!${CFV[col]}$103", "0.0000", "key",
     note="현물단가 + 조달이자 − 쿠폰FV")
brow(37, "▶ 선도 수익률", "%", lambda col, i: f"=현금흐름!${CFV[col]}$111*100", N_YLD, "key")
brow(38, "수렴 잔차 (0에 근접해야 정상)", "", lambda col, i: f"=현금흐름!${CFV[col]}$113", "0.00E+00")
brow(39, "선도 BPV", "포인트/1bp", lambda col, i: f"=현금흐름!${CFV[col]}$116", "0.000000")
brow(40, "선도 BPV (액면 1억 기준)", "원/1bp", lambda col, i: f"={col}39*설정!$B$9", N_WON2)
brow(41, "선도 2차미분 P''", "", lambda col, i: f"=현금흐름!${CFV[col]}$117", "0.00")
brow(42, "선도 잔존만기", "년", lambda col, i: f"=현금흐름!${CFV[col]}$118", "0.0000")
brow(43, "캐리 (선도−현물)", "bp", lambda col, i: f"=({col}37-{col}16)*100", N_BP,
     note="양수 = 선도수익률이 높음 = 네거티브 캐리")

section(bs, 45, "[헤지 / 포지션]", 7)
brow(46, "이론 액면 (선물 1계약당)", "원", lambda col, i:
     f"=IFERROR({col}10*(이론가!$B$18/{col}39)*설정!$B$8,0)", N_WON, "key", total="=SUM(C46:E46)",
     note="w × (표준물BPV / 선도BPV) × 거래단위  → 종목별 수익률 리스크 완전중립")
brow(47, "이론 액면 총계", "원", lambda col, i: f"={col}46*설정!$B$40", N_WON, total="=SUM(C47:E47)")
brow(48, "실제 보유액면", "원", 0, N_WON, "in", total="=SUM(C48:E48)")
brow(49, "▶ 적용 액면", "원", lambda col, i: f'=IF(설정!$B$45="예",{col}47,{col}48)', N_WON, "key",
     total="=SUM(C49:E49)")
brow(50, "진입 결제금액 (실제 체결)", "원", None, N_WON, "in", total="=SUM(C50:E50)",
     note="미입력 시 현재 단가 기준으로 자동 산출")
brow(51, "▶ 적용 진입금액", "원", lambda col, i: f'=IF({col}50="",{col}25/100*{col}49,{col}50)', N_WON, "key",
     total="=SUM(C51:E51)")
brow(52, "진입 환산단가 (더티)", "", lambda col, i: f'=IFERROR({col}51/{col}49*100,0)', "0.0000")
brow(53, "현재 평가금액", "원", lambda col, i: f"={col}25/100*{col}49", N_WON, total="=SUM(C53:E53)")
brow(54, "현재 평가손익 (방향반영)", "원", lambda col, i: f"=({col}53-{col}51)*설정!$B$39", N_WON,
     total="=SUM(C54:E54)")

bs["A56"] = ("※ 46행 헤지액면: 각 종목의 수익률 1bp 변동이 선물 leg와 정확히 상쇄되도록 산출합니다. "
             "가중치 합이 1이고 BPV가 비슷하면 3종목 합계는 대략 계약수 × 1억원이 됩니다.")
bs["A56"].font = F_NOTE
bs["A57"] = "※ 종목을 2개만 쓸 경우: 미사용 종목의 가중치(10행)를 0으로 두되, 발행일/만기일/표면금리는 유효한 값을 유지해야 오류가 없습니다."
bs["A57"].font = F_NOTE
bs["A58"] = "※ 4~8행 종목정보 출처: 한국거래소가 결제월별로 공시하는 최종결제기준채권. 현재 값은 형식 예시용 더미입니다."
bs["A58"].font = F_NOTE
bs.freeze_panes = "C4"

# ════════════════════════════════════════════════════════════════
# 5. 이론가
# ════════════════════════════════════════════════════════════════
th = wb.create_sheet("이론가")
title(th, "이론 선물가격 · 베이시스 · 차익 판정", 6)
for col, w in zip("ABCDEF", [34, 18, 12, 56, 10, 10]):
    th.column_dimensions[col].width = w


def stdprice(ycell):
    """표준물(5% 쿠폰, 반기) 폐쇄형 가격식"""
    return (f"=(설정!$B$6*100/({ycell}))*(1-(1/(1+({ycell})/2))^$B$12)"
            f"+100*(1/(1+({ycell})/2))^$B$12")


section(th, 3, "[1] 바스켓 선도수익률", 6)
th["C3"] = ""
for i, col in enumerate("CDE"):
    put(th, 4 + i, f"종목{i+1} 선도수익률", f"=바스켓!{col}37", "%", N_YLD, "link")
    th.cell(row=4 + i, column=3, value=f"=바스켓!{col}10").number_format = "0.0000"
    th.cell(row=4 + i, column=3).font = F_LINK
th["C3"] = "가중치"
th["C3"].font = F_HDR
put(th, 7, "▶ 가중평균 선도수익률", "=SUMPRODUCT(B4:B6,C4:C6)/SUM(C4:C6)", "%", N_YLD, "key")
put(th, 8, "(소수)", "=B7/100", "", "0.00000000")

section(th, 10, "[2] 표준물 가격 = 최종결제기준가격 산식", 6)
put(th, 11, "표준물 만기", "=설정!$B$5", "년", N_INT, "link")
put(th, 12, "반기 기간수 N", "=B11*설정!$B$7", "", N_INT)
put(th, 13, "반기 쿠폰", "=설정!$B$6*100/설정!$B$7", "", N_PX)
put(th, 14, "▶ 이론 선물가격", stdprice("$B$8"), "", "0.0000", "key")
put(th, 15, "이론가 (호가단위 반올림)", "=ROUND(B14/설정!$B$10,0)*설정!$B$10", "", N_PX2)
put(th, 16, "P(ȳ+1bp)", stdprice("$B$8+0.0001"), "", "0.000000")
put(th, 17, "P(ȳ−1bp)", stdprice("$B$8-0.0001"), "", "0.000000")
put(th, 18, "▶ 표준물 BPV", "=(B17-B16)/2", "포인트/1bp", "0.000000", "key")
put(th, 19, "표준물 BPV (원/1bp/계약)", "=B18*설정!$B$9", "원", N_WON2)
put(th, 20, "표준물 수정듀레이션", "=B18*10000/B14", "년", "0.0000")

section(th, 22, "[3] 시장 선물가격 → 내재수익률 (뉴턴 2회)", 6)
put(th, 23, "시장 선물가격", "=실시간!C5", "", N_PX2, "link")
put(th, 24, "ya (초기값)", "=B8", "", "0.00000000")
put(th, 25, "P(ya)", "=B14", "", "0.000000")
put(th, 26, "|dP/dy|", "=B18*10000", "", "0.000000")
put(th, 27, "yb", "=B24+(B25-B23)/B26", "", "0.00000000")
put(th, 28, "P(yb)", stdprice("$B$27"), "", "0.000000")
put(th, 29, "yc", "=B27+(B28-B23)/B26", "", "0.00000000")
put(th, 30, "P(yc)", stdprice("$B$29"), "", "0.000000")
put(th, 31, "▶ 선물 내재수익률", "=B29*100", "%", N_YLD, "key")
put(th, 32, "수렴 잔차", "=B30-B23", "", "0.00E+00")

section(th, 34, "[4] 베이시스 및 차익 판정", 6)
put(th, 35, "▶ 베이시스 (시장가 − 이론가)", "=B23-B14", "포인트", "0.0000", "key")
put(th, 36, "베이시스", "=B35/설정!$B$10", "틱", "0.00")
put(th, 37, "▶ 베이시스", "=B35*설정!$B$9", "원/계약", N_WON, "key")
put(th, 38, "수익률 스프레드 (선도평균 − 내재)", "=(B7-B31)*100", "bp", N_BP, "calc",
    "양수 = 선물 고평가")
put(th, 39, "총 거래비용", "=설정!$B$53", "원/계약", N_WON, "link")
put(th, 40, "▶ 순차익 (비용 차감후)", "=ABS(B37)-B39", "원/계약", N_WON, "key")
put(th, 41, "계약수", "=설정!$B$40", "계약", N_INT, "link")
put(th, 42, "▶ 총 순차익", "=B40*B41", "원", N_WON, "key")
put(th, 43, "권장 포지션",
    '=IF(B37>0,"선물 매도 + 바스켓 매수 (선물 고평가)",IF(B37<0,"선물 매수 + 바스켓 매도 (선물 저평가)","중립"))',
    "", None, "key")
put(th, 44, "설정된 방향과 일치?",
    '=IF(OR(AND(B37>0,설정!$B$39=1),AND(B37<0,설정!$B$39=-1)),"○ 일치","✕ 설정!B38 방향 확인")',
    "", None, "key")
put(th, 45, "판정", '=IF(B40>0,"차익거래 가능","비용 미달 — 관망")', "", None, "key")

th["A47"] = "※ 베이시스가 양수(선물 고평가)면 선물 매도 + 바스켓 매수(cash-and-carry), 음수면 반대(reverse cash-and-carry) 입니다."
th["A47"].font = F_NOTE
th["A48"] = "※ 40행은 절대값 기준입니다. 실제 진입 시에는 [설정] B38의 방향이 43행 권장과 일치하는지 44행에서 확인하십시오."
th["A48"].font = F_NOTE

# ════════════════════════════════════════════════════════════════
# 6. 손익
# ════════════════════════════════════════════════════════════════
pl = wb.create_sheet("손익")
title(pl, "만기 손익  ―  바스켓 1단위 = 선물 1계약 대응", 6)
for col, w in zip("ABCDEF", [40, 20, 12, 56, 10, 10]):
    pl.column_dimensions[col].width = w

section(pl, 3, "[A] 실시간 락인 손익 (지금 진입 → 만기 보유 가정)", 6)
put(pl, 4, "베이시스", "=이론가!B37", "원/계약", N_WON, "link")
put(pl, 5, "총 거래비용", "=설정!$B$53", "원/계약", N_WON, "link")
put(pl, 6, "▶ 순차익", "=이론가!B40", "원/계약", N_WON, "key")
put(pl, 7, "계약수 Q", "=설정!$B$40", "계약", N_INT, "link")
put(pl, 8, "▶ 총 순차익", "=B6*B7", "원", N_WON, "key")
put(pl, 9, "바스켓 투자금액 (현물 매입액)", "=SUM(바스켓!C53:E53)", "원", N_WON, "link")
put(pl, 10, "보유일수", "=설정!$B$28", "일", N_INT, "link")
put(pl, 11, "▶ 연환산 수익률", '=IF(AND(B9>0,B10>0),B8/B9*365/B10*100,"")', "%", "0.0000", "key")
put(pl, 12, "조달금리 대비 초과", '=IF(B11="","",B11-설정!$B$31)', "%p", "0.0000", "key",
    "레포로 100% 조달 시 이 값이 순마진")

section(pl, 14, "[B] 현재 미실현 손익 (보유 포지션 마킹)", 6)
put(pl, 15, "선물 진입가", "=설정!$B$42", "", N_PX2, "link")
put(pl, 16, "선물 현재가", "=실시간!C5", "", N_PX2, "link")
put(pl, 17, "선물 평가손익", "=(B15-B16)*설정!$B$39*설정!$B$9*설정!$B$40", "원", N_WON)
put(pl, 18, "현물 평가손익", "=바스켓!F54", "원", N_WON, "link")
put(pl, 19, "진입 경과일수", "=MAX(0,설정!$B$27-설정!$B$44)", "일", N_INT)
put(pl, 20, "경과 조달이자", "=-바스켓!F51*설정!$B$32*B19/365*설정!$B$39", "원", N_WON)
put(pl, 21, "▶ 미실현 손익 합계", "=B17+B18+B20", "원", N_WON, "key")

section(pl, 23, "[C] 만기 손익 — 시나리오", 6)
put(pl, 24, "만기 수익률 평행이동", 0, "bp", N_BP, "in", "만기시점 바스켓 수익률 shift")
put(pl, 25, "만기 바스켓 평균수익률", "=이론가!B7+B24/100", "%", N_YLD)
put(pl, 26, "▶ 최종결제기준가격", stdprice("$B$25/100").replace("$B$12", "이론가!$B$12"), "", "0.0000", "key")
for i, col in enumerate("CDE"):
    put(pl, 27 + i, f"종목{i+1} 만기 단가(더티)",
        f"=현금흐름!${CFV[col]}$121", "", "0.0000", "calc",
        "이표 스케줄 전체를 재할인한 정밀 재평가값 (근사 없음)")

pl["A31"] = "── C-1. 현재 시장가로 신규 진입한다고 가정 ──"
pl["A31"].font = Font(name=FONT, size=10, bold=True, color="1F3864")
put(pl, 32, "선물 leg", "=(이론가!B23-B26)*설정!$B$39*설정!$B$9*설정!$B$40", "원", N_WON)
put(pl, 33, "현물 leg",
    "=설정!$B$39*(($B$27-바스켓!C36)/100*바스켓!C49"
    "+($B$28-바스켓!D36)/100*바스켓!D49"
    "+($B$29-바스켓!E36)/100*바스켓!E49)", "원", N_WON,
    note="선도 목표단가 대비 처분손익. 조달이자·쿠폰FV는 선도단가에 이미 반영됨")
put(pl, 34, "거래비용", "=-설정!$B$53*설정!$B$40", "원", N_WON)
put(pl, 35, "▶ 순 만기손익", "=B32+B33+B34", "원", N_WON, "key")
put(pl, 36, "▶ 바스켓 1단위당 (=계약당)", '=IFERROR(B35/설정!$B$40,0)', "원", N_WON, "key")
put(pl, 37, "▶ 액면 1억원당", "=B36", "원", N_WON, "key")

pl["A39"] = "── C-2. 실제 보유 포지션 기준 ──"
pl["A39"].font = Font(name=FONT, size=10, bold=True, color="1F3864")
put(pl, 40, "선물 leg", "=(설정!$B$42-B26)*설정!$B$39*설정!$B$9*설정!$B$40", "원", N_WON)
put(pl, 41, "현물 취득금액", "=바스켓!F51", "원", N_WON, "link")
put(pl, 42, "조달일수 (진입일→H)", "=MAX(0,설정!$B$21-설정!$B$44)", "일", N_INT)
put(pl, 43, "조달 원리금 상환", "=-B41*(1+설정!$B$32*B42/365)", "원", N_WON)
put(pl, 44, "만기 처분금액",
    "=$B$27/100*바스켓!C49+$B$28/100*바스켓!D49+$B$29/100*바스켓!E49", "원", N_WON)
put(pl, 45, "쿠폰 재투자 수입",
    "=바스켓!C34/100*바스켓!C49+바스켓!D34/100*바스켓!D49+바스켓!E34/100*바스켓!E49", "원", N_WON)
put(pl, 46, "현물 leg", "=(B44+B45+B43)*설정!$B$39", "원", N_WON)
put(pl, 47, "거래비용", "=-설정!$B$53*설정!$B$40", "원", N_WON)
put(pl, 48, "▶ 순 만기손익", "=B40+B46+B47", "원", N_WON, "key")
put(pl, 49, "▶ 바스켓 1단위당", '=IFERROR(B48/설정!$B$40,0)', "원", N_WON, "key")

pl["A51"] = ("※ C-1은 '지금 이 순간 진입하면 만기에 얼마를 확정하는가' 이고, 헤지가 정확하면 B24(수익률 shift)를 "
             "바꿔도 B35가 거의 변하지 않아야 합니다 ([시나리오] 시트에서 확인).")
pl["A51"].font = F_NOTE
pl["A52"] = "※ C-2는 이미 체결한 가격(설정!B41, 바스켓!50행)을 기준으로 하며, 진입 이후 시장 변동으로 인한 실현 차이가 반영됩니다."
pl["A52"].font = F_NOTE
pl["A53"] = ("※ 27~29행 만기단가는 [현금흐름] 시트에서 이표 스케줄 전체를 재할인한 정밀 재평가값입니다(근사 없음). "
             "B24를 바꾸면 세 종목 모두 자동 재계산됩니다.")
pl["A53"].font = F_NOTE

# ════════════════════════════════════════════════════════════════
# 7. 시나리오
# ════════════════════════════════════════════════════════════════
sc = wb.create_sheet("시나리오")
title(sc, "만기 손익 민감도  ―  바스켓 수익률 평행이동 (헤지 유효성 점검)", 9)
for col, w in zip("ABCDEFGHI", [12, 16, 16, 18, 18, 16, 18, 20, 6]):
    sc.column_dimensions[col].width = w

heads = ["shift (bp)", "만기 평균수익률(%)", "최종결제가격", "선물 leg (원)",
         "현물 leg (원)", "거래비용 (원)", "순 만기손익 (원)", "바스켓 1단위당 (원)"]
for i, h in enumerate(heads):
    c = sc.cell(row=3, column=1 + i, value=h)
    c.font, c.fill, c.border = F_HDR, FILL_HDR, BOX
    c.alignment = Alignment(horizontal="center", wrap_text=True)

R0 = 4
for i in range(21):
    R = R0 + i
    shift = -100 + i * 10
    sc.cell(row=R, column=1, value=shift).number_format = N_BP
    sc.cell(row=R, column=1).font = F_IN
    sc.cell(row=R, column=2, value=f"=이론가!$B$7+$A{R}/100").number_format = N_YLD
    sc.cell(row=R, column=3,
            value=f"=(설정!$B$6*100/($B{R}/100))*(1-(1/(1+($B{R}/100)/2))^이론가!$B$12)"
                  f"+100*(1/(1+($B{R}/100)/2))^이론가!$B$12").number_format = "0.0000"
    sc.cell(row=R, column=4,
            value=f"=(이론가!$B$23-$C{R})*설정!$B$39*설정!$B$9*설정!$B$40").number_format = N_WON
    _pr = 123 + i   # 현금흐름 시트의 시나리오별 정밀 재평가 행
    sc.cell(row=R, column=5,
            value=f"=설정!$B$39*("
                  f"(현금흐름!$B${_pr}-바스켓!C36)/100*바스켓!C49"
                  f"+(현금흐름!$AR${_pr}-바스켓!D36)/100*바스켓!D49"
                  f"+(현금흐름!$CH${_pr}-바스켓!E36)/100*바스켓!E49)").number_format = N_WON
    sc.cell(row=R, column=6, value="=-설정!$B$53*설정!$B$40").number_format = N_WON
    sc.cell(row=R, column=7, value=f"=$D{R}+$E{R}+$F{R}").number_format = N_WON
    sc.cell(row=R, column=8, value=f"=IFERROR($G{R}/설정!$B$40,0)").number_format = N_WON
    for cc in range(1, 9):
        cell = sc.cell(row=R, column=cc)
        cell.border = BOX
        if cc > 1:
            cell.font = F_CALC
    if shift == 0:
        for cc in range(1, 9):
            sc.cell(row=R, column=cc).fill = FILL_KEY

put(sc, 27, "손익 최대값", "=MAX(G4:G24)", "원", N_WON, "key", lab_col=1, val_col=2, unit_col=3, note_col=4)
put(sc, 28, "손익 최소값", "=MIN(G4:G24)", "원", N_WON, "key", lab_col=1, val_col=2, unit_col=3, note_col=4)
put(sc, 29, "변동폭 (헤지 잔여리스크)", "=B27-B28", "원", N_WON, "key", lab_col=1, val_col=2, unit_col=3, note_col=4)
put(sc, 30, "변동폭 ÷ shift0 손익", '=IFERROR(B29/ABS(G14),0)', "배", "0.0000", "key",
    lab_col=1, val_col=2, unit_col=3, note_col=4)
put(sc, 31, "shift 0 기준 손익", "=G14", "원", N_WON, "key",
    lab_col=1, val_col=2, unit_col=3, note_col=4)
sc["A33"] = ("※ BPV 헤지는 1차(듀레이션) 리스크만 중립화합니다. 위 표가 shift 0에서 볼록하게 꺾여 내려가는 것은 "
             "이 거래가 구조적으로 '컨벡시티 매도' 포지션이기 때문입니다 — 표준물(5% 쿠폰)이 바스켓(저쿠폰)보다 "
             "컨벡시티가 커서, 선물을 매도하면 컨벡시티를 파는 셈이 됩니다. 헤지로 제거되지 않습니다.")
sc["A33"].font = F_NOTE
sc["A34"] = ("※ 평행이동만 가정합니다. 바스켓 종목 간 상대수익률(커브) 변화는 종목별 BPV 헤지로 이미 중립화되어 있습니다.")
sc["A35"] = ("※ 현물 leg는 근사가 아니라 이표 스케줄 전체를 재할인한 정밀 재평가입니다. A열의 shift 값을 바꾸면 "
             "[현금흐름] 시트가 자동으로 해당 수익률에서 재평가합니다.")
sc["A35"].font = F_NOTE
sc["A34"].font = F_NOTE
sc.freeze_panes = "A4"

# ════════════════════════════════════════════════════════════════
# 8. 연동가이드
# ════════════════════════════════════════════════════════════════
gd = wb.create_sheet("연동가이드")
title(gd, "실시간 데이터 연동 가이드 & 준비물 체크리스트", 5)
gd.column_dimensions["A"].width = 4
gd.column_dimensions["B"].width = 26
gd.column_dimensions["C"].width = 92

rows = [
    ("s", "A. 벤더별 실시간 연동 수식", ""),
    ("h", "벤더 / 방식", "[실시간] C열에 넣을 수식 예시"),
    ("r", "연합인포맥스 (Excel Add-in)", '=IMFX("KR103502GE97","매도수익률")   /   =IMFX("F 10YKTB","현재가")'),
    ("r", "Bloomberg (BDP)", '=BDP("KTBA Comdty","PX_LAST")   /   =BDP("KTB 3 06/10/35 Corp","YLD_YTM_ASK")'),
    ("r", "Bloomberg (RTD 스트리밍)", '=RTD("bloomberglp.bloombergrtd",,"//blp/mktdata","KTBA Comdty","LAST_PRICE")'),
    ("r", "Refinitiv Eikon", '=RtGet("IDN","KTBc1","CF_LAST")   또는  =TR("KTBc1","CF_LAST")'),
    ("r", "LS증권 xingAPI (DDE)", '=xingDDE|주식현재가!\'종목코드\'   (xingAPI DDE 서버 실행 필요)'),
    ("r", "대신증권 CYBOS Plus (DDE)", '=CYBOS|StockMst!\'A005930\'   (CYBOS Plus 로그인 상태 필요)'),
    ("r", "키움 OpenAPI+", 'Excel 직결 미지원 → VBA/파이썬 브리지로 셀에 write, 또는 RTD 서버 자작'),
    ("r", "한국투자증권 OpenAPI (REST)", 'Power Query 또는 VBA로 폴링. 실시간은 WebSocket → 별도 브리지 필요'),
    ("r", "KRX 정보데이터시스템", '지연 데이터. 차익거래 실행용으로는 부적합(참고용)'),
    ("n", "", "※ 벤더 함수명·필드명은 버전에 따라 다릅니다. 반드시 해당 Add-in 매뉴얼로 확인하십시오."),
    ("n", "", "※ RTD/DDE는 [파일]-[옵션]-[수식]에서 계산옵션이 '자동'이어야 갱신됩니다."),
    ("s", "B. 준비물 체크리스트 — 이 파일을 실전에 쓰려면 아래가 추가로 필요합니다", ""),
    ("h", "항목", "내용"),
    ("r", "1. 바스켓 종목 확정", "해당 결제월의 최종결제기준채권(종목코드/발행일/만기일/표면금리). KRX가 결제월별로 공시."),
    ("r", "2. 실시간 채권 호가", "바스켓 종목의 Bid/Ask 수익률. 장외 채권은 유동성이 얕아 '체결 가능한' 호가인지가 관건."),
    ("r", "3. 실시간 선물 호가", "최근월물 현재가 + 1호가. 체결가 기준으로 베이시스를 봐야 실행 가능."),
    ("r", "4. 조달금리(레포)", "선물 만기까지의 term repo 금리. 담보 haircut, 롤 가능 기간, 종목별 special 여부."),
    ("r", "5. 현물 조달·대차 가능성", "역방향(선물매수+현물매도) 시 바스켓 종목 대차 가능 여부와 대차수수료."),
    ("r", "6. 거래비용 실측치", "선물 위탁수수료(계약당), 채권 장외 스프레드, 결제·최종결제 슬리피지."),
    ("r", "7. 증거금 / 자금계획", "선물 개시·유지증거금, 일일정산 변동증거금 조달. 마진콜 시나리오."),
    ("r", "8. 휴일 캘린더", "최종결제일·현물결제일·이표일의 영업일 조정. 현재는 WORKDAY() 기반이라 공휴일 미반영."),
    ("r", "9. 세무 / 회계 처리", "기관 유형별 채권 이자소득 원천징수 여부 → 실효 캐리에 직접 영향."),
    ("r", "10. 최종결제 산출 규칙", "딜러 제출 수익률의 집계 시각·최고최저 제외 평균 여부 등 KRX 현행 규정 재확인."),
    ("s", "C. 남는 리스크 (헤지로 제거되지 않는 것)", ""),
    ("r", "결제가격 괴리 리스크", "최종결제기준가격은 딜러 호가 집계값. 만기일에 내가 실제 체결하는 현물 수익률과 다를 수 있음 → 최대 리스크."),
    ("r", "레포 롤 리스크", "만기까지 term repo를 못 잡으면 조달금리가 변동 → 캐리 전제가 무너짐."),
    ("r", "컨벡시티 불일치", "BPV 헤지는 1차만 중립화. 대폭 금리변동 시 [시나리오] 시트의 잔여 변동만큼 손익 흔들림."),
    ("r", "변동증거금 이자효과", "선물은 일일정산되나 현물은 만기 일시정산 → 금리·손익 경로에 따른 테일 리스크."),
    ("r", "유동성 리스크", "바스켓 종목이 비지표물이면 진입·청산 스프레드가 이론 베이시스를 잠식."),
]
r = 3
for kind, a, b in rows:
    if kind == "s":
        gd.cell(row=r, column=2, value=a).font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        for cc in (2, 3):
            gd.cell(row=r, column=cc).fill = FILL_SEC
        gd.cell(row=r, column=3).font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    elif kind == "h":
        for cc, txt in ((2, a), (3, b)):
            c = gd.cell(row=r, column=cc, value=txt)
            c.font, c.fill, c.border = F_HDR, FILL_HDR, BOX
    elif kind == "n":
        as_text(gd.cell(row=r, column=3, value=b)).font = F_NOTE
    else:
        gd.cell(row=r, column=2, value=a).font = F_LBL
        gd.cell(row=r, column=2).border = BOX
        c = as_text(gd.cell(row=r, column=3, value=b))
        c.font, c.border = F_LBL, BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# 시트 순서 정리
wb.move_sheet("README", offset=0)
wb._sheets = [wb["README"], wb["실시간"], wb["설정"], wb["바스켓"],
              wb["이론가"], wb["손익"], wb["시나리오"], wb["현금흐름"], wb["연동가이드"]]

# 최종 안전망: 설명·예시 영역에 남은 '='로 시작하는 문자열을 텍스트로 강제
for _ws, _cols in ((lv, (4, 5)), (gd, (2, 3)), (bs, (7,)), (cf, (4,)),
                   (th, (4,)), (pl, (4,)), (wb["README"], (2,))):
    for _row in _ws.iter_rows():
        for _c in _row:
            if _c.column in _cols:
                as_text(_c)

wb.save(OUT)
print("saved:", OUT)
