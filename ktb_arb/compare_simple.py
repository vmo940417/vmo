# -*- coding: utf-8 -*-
"""
간소판 검증: IMDP(인포맥스) 셀만 숫자로 치환한 사본을 formulas 로 평가하고
파이썬 독립 모델과 대조한다.
"""
import datetime as dt
import os
import shutil
from pathlib import Path
from dateutil.relativedelta import relativedelta as rd
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
XL = HERE / "국채선물_차익거래_간소.xlsx"
TMP = HERE / f"_간소_eval_{os.getpid()}.xlsx"

# 원본 xlsm 캐시값 (실제 시장 수준)
LIVE = {"B25": 105.23, "B27": 4.309, "B28": 4.335, "B29": 105.30}


def ed(d, m):
    return d + rd(months=m)


def kr_price(y, settle, issue, mat, cpn, freq=2):
    cfs, k = [], 0
    while True:
        d = ed(mat, -6 * k)
        if d <= issue:
            break
        cfs.append((d, cpn * 100 / freq + (100 if d == mat else 0)))
        k += 1
    cfs = sorted([x for x in cfs if x[0] > settle])
    nxt = cfs[0][0]
    prv = ed(nxt, -6)
    dd, T = (nxt - settle).days, (nxt - prv).days
    return sum(a / (1 + y / 2) ** i for i, (_, a) in enumerate(cfs)) / (1 + y / 2 * dd / T)


def std_price(y, n, cpn=0.05):
    v = 1 / (1 + y / 2)
    N = 2 * n
    return (cpn * 100 / y) * (1 - v ** N) + 100 * v ** N


def fwd_2step(y, S0, H, issue, mat, cpn, r):
    """워크북과 동일한 뉴턴 2회 + y1 기준 BPV. 정확해(이분법)도 함께 반환."""
    P0 = kr_price(y, S0, issue, mat, cpn)
    fv, k = 0.0, 0
    while True:
        d = ed(mat, -6 * k)
        if d <= issue:
            break
        if S0 < d <= H:
            fv += (cpn * 100 / 2) * (1 + r * (H - d).days / 365)
        k += 1
    tgt = P0 + P0 * r * (H - S0).days / 365 - fv
    PH0 = kr_price(y, H, issue, mat, cpn)
    slope = (kr_price(y - 1e-4, H, issue, mat, cpn)
             - kr_price(y + 1e-4, H, issue, mat, cpn)) / 2 * 10000
    y1 = y + (PH0 - tgt) / slope
    bpv = (kr_price(y1 - 1e-4, H, issue, mat, cpn)
           - kr_price(y1 + 1e-4, H, issue, mat, cpn)) / 2
    y2 = y1 + (kr_price(y1, H, issue, mat, cpn) - tgt) / slope
    lo, hi = -0.5, 1.0
    for _ in range(300):
        mid = (lo + hi) / 2
        if kr_price(mid, H, issue, mat, cpn) > tgt:
            lo = mid
        else:
            hi = mid
    return P0, tgt, y2, bpv, (lo + hi) / 2


def evaluate(extra=None):
    shutil.copy(XL, TMP)
    wb = load_workbook(TMP)
    for coord, val in {**LIVE, **(extra or {})}.items():
        wb["메인"][coord] = val
    wb.save(TMP)
    import warnings
    warnings.filterwarnings("ignore")
    import formulas
    sol = formulas.ExcelModel().loads(str(TMP)).finish().calculate()
    out = {}
    for k, v in sol.items():
        try:
            val = v.value[0, 0]
        except Exception:
            try:
                val = v.value
            except Exception:
                val = v
        try:
            hash(val)
        except Exception:
            val = str(val)
        out[k.upper().replace("'", "")] = val
    return out


def main():
    sol = evaluate()
    book = TMP.name.upper()

    def G(sheet, coord):
        key = f"[{book}]{sheet.upper()}!{coord}"
        if key in sol:
            return sol[key]
        for k, v in sol.items():
            if k.endswith(f"]{sheet.upper()}!{coord}"):
                return v
        return None

    errs = [(k, str(v)) for k, v in sol.items()
            if any(e in str(v) for e in ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!",
                                         "#N/A", "#NUM!", "#NULL!"))]
    print(f"■ 오류값 셀: {len(errs)}개")
    for k, v in errs[:20]:
        print("   ", k, "=", v)
    print()

    def to_date(x):
        if isinstance(x, (int, float)):
            return dt.date(1899, 12, 30) + dt.timedelta(days=float(x))
        return x.date() if isinstance(x, dt.datetime) else x

    S0, H = to_date(G("메인", "B10")), to_date(G("메인", "B8"))
    r = float(G("메인", "B31")) / 100
    n_std = int(float(G("메인", "B14")))
    mult = float(G("메인", "B16"))
    Q = float(G("메인", "B17"))
    print(f"S0={S0}  H={H}  보유일수={(H-S0).days}  조달={r:.4%}  "
          f"표준물={n_std}년  1pt={mult:,.0f}원  Q={Q:.0f}\n")

    wf = load_workbook(XL)
    mm = wf["메인"]
    ok = True
    fw, ws_, tgts, bpvs, infos = [], [], [], [], []
    for col in "CD":
        issue = to_date(mm[f"{col}37"].value)
        mat = to_date(mm[f"{col}38"].value)
        cpn = mm[f"{col}36"].value / 100
        y = float(G("메인", f"{col}41"))
        w = float(G("메인", f"{col}39"))
        P0, tgt, y1, bpv, y_ex = fwd_2step(y, S0, H, issue, mat, cpn, r)
        rows = [("현물단가", P0, float(G("메인", f"{col}44")), 1e-6),
                ("선도목표단가", tgt, float(G("메인", f"{col}49")), 1e-6),
                ("선도수익률%", y1 * 100, float(G("메인", f"{col}50")), 1e-7),
                ("선도BPV", bpv, float(G("메인", f"{col}52")), 1e-8)]
        for nm, a, b, tol in rows:
            d = abs(a - b)
            if d >= tol:
                ok = False
            print(f"[{col}] {nm:13s} py={a:>14.9f}  xl={b:>14.9f}  "
                  f"diff={d:.2e}  {'OK' if d < tol else '★FAIL★'}")
        # 이표일 자동계산 확인
        xl_next = to_date(G("메인", f"{col}42"))
        py_next = min(d for d in (ed(mat, -6 * k) for k in range(60)) if d > S0)
        print(f"[{col}] 차기이표일    py={py_next}  xl={xl_next}  "
              f"{'OK' if py_next == xl_next else '★FAIL★'}")
        if py_next != xl_next:
            ok = False
        resid = kr_price(y1, H, issue, mat, cpn) - tgt
        print(f"[{col}] 정확해 대비   뉴턴2회={y1*100:.9f}%  이분법={y_ex*100:.9f}%  "
              f"오차={abs(y1-y_ex)*1e4:.2e}bp")
        print(f"[{col}] shift0 누수  잔차={resid:+.3e} → "
              f"{resid/100*float(G('메인', f'{col}58')):>10,.0f}원  "
              f"(엑셀 표시 잔차 {float(G('메인', f'{col}51')):.3e})\n")
        fw.append(y1); ws_.append(w); tgts.append(tgt); bpvs.append(bpv)
        infos.append((issue, mat, cpn, float(G("메인", f"{col}50")) / 100,
                      float(G("메인", f"{col}49"))))

    ybar = sum(a * b for a, b in zip(fw, ws_)) / sum(ws_)
    pth = std_price(ybar, n_std)
    print(f"평균 선도수익률   py={ybar*100:.8f}%  xl={float(G('메인','B64')):.8f}%")
    print(f"이론 선물가격     py={pth:.8f}     xl={float(G('메인','B66')):.8f}  "
          f"diff={abs(pth-float(G('메인','B66'))):.2e}")
    if abs(pth - float(G("메인", "B66"))) > 1e-6:
        ok = False
    std_bpv = (std_price(ybar - 1e-4, n_std) - std_price(ybar + 1e-4, n_std)) / 2
    print(f"표준물 BPV        py={std_bpv:.8f}     xl={float(G('메인','B60')):.8f}")
    print(f"베이시스          xl={float(G('메인','B69')):,.0f}원/계약  "
          f"(시장 {LIVE['B25']} − 이론 {float(G('메인','B66')):.4f})")
    print(f"[참고] 현물금리 기준 이론가 xl={float(G('메인','B77')):.4f}  "
          f"→ 캐리 {float(G('메인','B78')):,.0f}원/계약\n")

    print("헤지액면 (계약 1개당):")
    tot = 0
    for col in "CD":
        vv = float(G("메인", f"{col}55"))
        tot += vv
        print(f"   [{col}] {vv:>16,.0f}원")
    print(f"   합계  {tot:>16,.0f}원\n")

    print("만기손익:")
    for c, lab in [("B85", "최종결제가"), ("B86", "선물 leg"), ("B87", "현물 leg"),
                   ("B88", "거래비용"), ("B89", "순 만기손익"), ("B90", "계약당")]:
        print(f"   메인!{c} {lab:11s} = {float(G('메인', c)):>16,.2f}")

    print("\n시나리오 — 엑셀 vs 파이썬 정밀 재할인:")
    Fm = float(G("메인", "B21"))
    cost = float(G("메인", "B22")) * Q
    faces = [float(G("메인", f"{c}58")) for c in "CD"]
    worst = 0.0
    for i in range(9):
        R = 96 + i
        sh = float(G("메인", f"A{R}"))
        xlv = float(G("메인", f"F{R}"))
        fut = (Fm - std_price(ybar + sh / 1e4, n_std)) * mult * Q
        cash = sum((kr_price(y1 + sh / 1e4, H, iss, mt, cp) - tg) / 100 * fc
                   for (iss, mt, cp, y1, tg), fc in zip(infos, faces))
        pyv = fut + cash - cost
        worst = max(worst, abs(pyv - xlv))
        print(f"   {sh:+6.0f}bp  xl={xlv:>15,.0f}  py={pyv:>15,.0f}  diff={pyv-xlv:>9,.2f}")
    print(f"   최대오차 = {worst:,.4f}원")
    if worst > 1.0:
        ok = False
    print(f"   변동폭(잔여리스크) = {float(G('메인','B108')):,.0f}원")

    TMP.unlink(missing_ok=True)

    # 방향 전환 시 금리가 바뀌는지 확인
    print("\n■ 포지션 방향별 적용금리")
    for d, fee, expect in (("선물매도+현물매수", 0.0, 2.60),
                           ("선물매수+현물매도", 0.0, 2.90),
                           ("선물매수+현물매도", 20.0, 2.70),
                           ("선물매도+현물매수", 20.0, 2.40)):
        s2 = evaluate({"B18": d, "B23": fee})
        idx2 = {k.upper().replace("'", ""): v for k, v in s2.items()}

        def G2(sheet, coord):
            for k, v in idx2.items():
                if k.endswith(f"]{sheet.upper()}!{coord}"):
                    return v
        got = float(G2("메인", "B31"))
        eng = float(G2("엔진", "B9"))
        sign = float(G2("메인", "B19"))
        good = abs(got - expect) < 1e-9 and abs(eng - expect / 100) < 1e-12
        if not good:
            ok = False
        print(f"   {d}  수수료={fee:>4.0f}bp  방향계수={sign:+.0f}  적용금리={got:.4f}%  "
              f"엔진!B9={eng:.6f}  기대 {expect:.2f}%  {'OK' if good else '★FAIL★'}")
        print(f"      → 선도수익률 종목1 {float(G2('메인','C50')):.6f}%  "
              f"종목2 {float(G2('메인','D50')):.6f}%   "
              f"이론가 {float(G2('메인','B66')):.4f}")
        TMP.unlink(missing_ok=True)

    print("\n=== 판정:", "PASS" if (ok and not errs) else "★확인 필요★", "===")


main()
