# -*- coding: utf-8 -*-
"""엑셀 결과 검증: 동일한 모형을 파이썬으로 독립 구현해 대조한다."""
import datetime as dt
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from pathlib import Path

XL = str(Path(__file__).resolve().parent / "국채선물_차익거래.xlsx")


def edate(d, months):
    return d + relativedelta(months=months)


def kr_price(y, settle, issue, maturity, cpn, freq=2):
    """한국 관행식 더티단가 (100 기준). y: 소수"""
    cfs = []
    k = 0
    while True:
        d = edate(maturity, -6 * k)
        if d <= issue:
            break
        amt = cpn * 100 / freq + (100 if d == maturity else 0)
        cfs.append((d, amt))
        k += 1
    cfs = sorted([c for c in cfs if c[0] > settle])
    nxt = cfs[0][0]
    prv = edate(nxt, -6)
    d_, T_ = (nxt - settle).days, (nxt - prv).days
    s = sum(amt / (1 + y / 2) ** i for i, (_, amt) in enumerate(cfs))
    return s / (1 + y / 2 * d_ / T_)


def std_price(y, n, cpn=0.05):
    v = 1 / (1 + y / 2)
    N = 2 * n
    return (cpn * 100 / y) * (1 - v ** N) + 100 * v ** N


def fwd_yield(y_spot, settle, H, issue, maturity, cpn, r, rr):
    P0 = kr_price(y_spot, settle, issue, maturity, cpn)
    # 기간중 쿠폰 재투자
    fv = 0.0
    k = 0
    while True:
        d = edate(maturity, -6 * k)
        if d <= issue:
            break
        if settle < d <= H:
            fv += (cpn * 100 / 2) * (1 + rr * (H - d).days / 365)
        k += 1
    target = P0 + P0 * r * (H - settle).days / 365 - fv
    # 이분법으로 정확해
    lo, hi = -0.5, 1.0
    for _ in range(200):
        mid = (lo + hi) / 2
        if kr_price(mid, H, issue, maturity, cpn) > target:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2, P0, target


def main():
    wv = load_workbook(XL, data_only=True)
    wf = load_workbook(XL)
    cfg, bsk, thr, pl, lv = (wv["설정"], wv["바스켓"], wv["이론가"], wv["손익"], wv["실시간"])

    settle = cfg["B27"].value
    H = cfg["B21"].value
    if isinstance(settle, dt.datetime):
        settle = settle.date()
    if isinstance(H, dt.datetime):
        H = H.date()
    r = cfg["B32"].value
    rr = cfg["B34"].value
    n_std = int(cfg["B5"].value)
    mult = cfg["B9"].value

    print(f"결제일 S0 = {settle},  최종결제일 H = {H},  보유일수 = {(H-settle).days}")
    print(f"조달금리 r = {r:.6f},  표준물 만기 = {n_std}년,  1포인트 = {mult:,.0f}원\n")

    ok = True
    fwds, ws_ = [], []
    for col in "CDE":
        issue = bsk[f"{col}6"].value
        mat = bsk[f"{col}7"].value
        if isinstance(issue, dt.datetime):
            issue = issue.date()
        if isinstance(mat, dt.datetime):
            mat = mat.date()
        cpn = bsk[f"{col}8"].value / 100
        y = bsk[f"{col}17"].value
        w = bsk[f"{col}10"].value

        py_fwd, py_P0, py_tgt = fwd_yield(y, settle, H, issue, mat, cpn, r, rr)
        xl_P0 = bsk[f"{col}25"].value
        xl_tgt = bsk[f"{col}36"].value
        xl_fwd = bsk[f"{col}37"].value / 100

        for name, a, b, tol in [("현물단가", py_P0, xl_P0, 1e-6),
                                ("선도목표단가", py_tgt, xl_tgt, 1e-6),
                                ("선도수익률", py_fwd, xl_fwd, 1e-9)]:
            diff = abs(a - b)
            flag = "OK " if diff < tol else "FAIL"
            if diff >= tol:
                ok = False
            print(f"[{col}] {name:12s} py={a:.10f}  xl={b:.10f}  diff={diff:.2e}  {flag}")
        fwds.append(py_fwd)
        ws_.append(w)
        print(f"[{col}] 수렴잔차(엑셀) = {bsk[f'{col}38'].value:.2e}")
        print()

    ybar = sum(f * w for f, w in zip(fwds, ws_)) / sum(ws_)
    py_theo = std_price(ybar, n_std)
    xl_theo = thr["B14"].value
    print(f"가중평균 선도수익률  py={ybar*100:.6f}%  xl={thr['B7'].value:.6f}%")
    print(f"이론 선물가격        py={py_theo:.6f}    xl={xl_theo:.6f}   diff={abs(py_theo-xl_theo):.2e}")
    if abs(py_theo - xl_theo) > 1e-6:
        ok = False

    fmkt = lv["C5"].value
    py_basis = (fmkt - py_theo) * mult
    print(f"베이시스(원/계약)    py={py_basis:,.0f}   xl={thr['B37'].value:,.0f}")
    print(f"선물 내재수익률      xl={thr['B31'].value:.6f}%  (잔차 {thr['B32'].value:.2e})")
    print(f"표준물 BPV           xl={thr['B18'].value:.6f} 포인트 = {thr['B19'].value:,.1f}원\n")

    print("헤지액면(선물 1계약당):")
    tot = 0
    for col in "CDE":
        v = bsk[f"{col}46"].value
        tot += v
        print(f"  [{col}] {v:>18,.0f}원   선도BPV={bsk[f'{col}39'].value:.6f}")
    print(f"  합계   {tot:>18,.0f}원   (≈ 1억이면 정상)\n")

    print("만기손익 [C-1] 시장가 신규진입 기준:")
    for a in ["B32", "B33", "B34", "B35", "B36"]:
        print(f"  손익!{a} = {pl[a].value:,.2f}")

    print("\n시나리오 헤지 유효성:")
    sc = wv["시나리오"]
    for row in [4, 9, 14, 19, 24]:
        print(f"  shift {sc[f'A{row}'].value:+6.0f}bp → 순손익 {sc[f'G{row}'].value:>15,.0f}원")
    print(f"  변동폭(잔여리스크) = {sc['B29'].value:,.0f}원")

    # 수식 셀 중 캐시값이 None인 것 탐지
    print("\n미평가(None) 수식 셀 점검:")
    bad = 0
    for name in wf.sheetnames:
        sf, sv = wf[name], wv[name]
        for row in sf.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    if sv[c.coordinate].value is None:
                        bad += 1
                        if bad <= 15:
                            print(f"  {name}!{c.coordinate}: {c.value[:70]}")
    print(f"  총 {bad}개")

    print("\n=== 판정:", "PASS" if ok and bad == 0 else "확인 필요", "===")


main()
