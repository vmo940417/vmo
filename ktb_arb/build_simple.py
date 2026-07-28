# -*- coding: utf-8 -*-
"""
국채선물(10년) 차익거래 — 간소화판 생성기

기존 KTB10 v2.xlsm 을 참조해 다음을 반영한 축소본:
  · 시트 2개 (메인 / 엔진)          ← 기존 자체판 9개
  · 바스켓 2종목 (국고 25-5, 25-11) ← 사용자 실제 구성
  · 연합인포맥스 _xll.IMDP 실수식    ← 사용자 실제 벤더
  · VBA 없음 (.xlsx)

원본 xlsm 대비 수정한 것:
  1. 이론가를 '선도'금리 기준으로 산출 → 캐리 이중계상 제거
  2. 이표일을 만기일에서 자동 역산 → 하드코딩 스테일 문제 해소
  3. 선물 듀레이션을 표준물(5% 쿠폰) 정식 BPV로 → 계약수 오차 제거
  4. 선도가격 수식의 텍스트 참조($F$2) 버그 제거
  5. 만기 손익 / 시나리오 추가 (원래 목표)

산출물: 국채선물_차익거래_간소.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
import datetime as dt

OUT = str(Path(__file__).resolve().parent / "국채선물_차익거래_간소.xlsx")

FONT = "Arial"
F_TITLE = Font(name=FONT, size=13, bold=True, color="FFFFFF")
F_SEC = Font(name=FONT, size=10, bold=True, color="FFFFFF")
F_HDR = Font(name=FONT, size=9, bold=True)
F_LBL = Font(name=FONT, size=9)
F_IN = Font(name=FONT, size=9, color="0000FF")
F_LIVE = Font(name=FONT, size=9, bold=True, color="C00000")
F_CALC = Font(name=FONT, size=9)
F_KEY = Font(name=FONT, size=10, bold=True)
F_NOTE = Font(name=FONT, size=8, italic=True, color="808080")

FILL_TITLE = PatternFill("solid", fgColor="1F3864")
FILL_SEC = PatternFill("solid", fgColor="4472C4")
FILL_IN = PatternFill("solid", fgColor="FFF2CC")
FILL_LIVE = PatternFill("solid", fgColor="DDEBF7")
FILL_KEY = PatternFill("solid", fgColor="E2EFDA")
FILL_HDR = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WON, PX, YLD, BP, DATE, INT_ = '#,##0', '0.0000', '0.0000', '0.00', 'yyyy-mm-dd', '#,##0'

# 시나리오 shift (bp) — 9개
SHIFTS = [-100, -75, -50, -25, 0, 25, 50, 75, 100]
NPER = 62
SCH0 = 15
SCH1 = SCH0 + NPER - 1          # 76
BLOCKS = [1, 31]                # 종목1: A~AB(1..28), 종목2: AE~BF(31..58)
NCOL = 28

def as_text(cell):
    """'='로 시작하는 설명·비고 문자열을 수식이 아닌 텍스트로 저장한다."""
    if isinstance(cell.value, str) and cell.value.startswith("="):
        cell.data_type = "s"
    return cell


wb = Workbook()

# ═══════════════════════════════════════════════════════════
# 메인
# ═══════════════════════════════════════════════════════════
m = wb.active
m.title = "메인"
for col, w in zip("ABCDEFGH", [30, 16, 16, 16, 10, 52, 10, 10]):
    m.column_dimensions[col].width = w
m["A1"] = "국채선물(10년) 차익거래 — 실시간 만기손익"
m["A1"].font, m["A1"].fill = F_TITLE, FILL_TITLE
for c in range(1, 9):
    m.cell(row=1, column=c).fill = FILL_TITLE
m.row_dimensions[1].height = 20


def sec(row, text):
    m.cell(row=row, column=1, value=text).font = F_SEC
    for c in range(1, 9):
        m.cell(row=row, column=c).fill = FILL_SEC


def put(row, label, value, unit="", fmt=None, style="calc", note=""):
    as_text(m.cell(row=row, column=1, value=label)).font = F_LBL
    c = m.cell(row=row, column=2, value=value)
    c.font = {"in": F_IN, "live": F_LIVE, "key": F_KEY}.get(style, F_CALC)
    fill = {"in": FILL_IN, "live": FILL_LIVE, "key": FILL_KEY}.get(style)
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    c.border = BOX
    if unit:
        as_text(m.cell(row=row, column=3, value=unit)).font = F_NOTE
    if note:
        as_text(m.cell(row=row, column=6, value=note)).font = F_NOTE
    return c


# ── [1] 설정 ────────────────────────────────────────────────
sec(3, "[1] 설정   ◆ 노란 셀만 입력하십시오")
put(4, "결제월 (YYYYMM)", 202609, "", INT_, "in", "3/6/9/12월물")
put(5, "최종거래일 (자동: 셋째 화요일)",
    "=DATE(INT(B4/100),MOD(B4,100),1)"
    "+MOD(2-WEEKDAY(DATE(INT(B4/100),MOD(B4,100),1),2)+7,7)+14", "", DATE)
put(6, "최종거래일 (수동 우선)", None, "", DATE, "in", "공휴일 조정 시 입력")
put(7, "▶ 최종거래일", '=IF(B6="",B5,B6)', "", DATE, "key")
put(8, "▶ 정산기준일 H", "=WORKDAY(B7,1)", "", DATE, "key", "현물 청산 T+1 결제일")
put(9, "현재일", "=TODAY()", "", DATE)
put(10, "▶ 현물결제일 S0", "=WORKDAY(B9,1)", "", DATE, "key", "국고채 익일결제")
put(11, "보유일수 (S0 → H)", "=B8-B10", "일", INT_, "key")
put(12, "조달금리 — 현물매수 시 지급", 2.60, "%", YLD, "in",
    "레포(RP) 조달금리. 선물매도+현물매수 방향에서 적용")
put(13, "담보운용금리 — 현물매도 시 수취", 2.90, "%", YLD, "in",
    "역RP 운용금리. 선물매수+현물매도 방향에서 적용")
put(14, "표준물 만기", 10, "년", INT_, "in", "10년물 고정")
put(15, "표준물 표면금리", 0.05, "", "0.00%", "in", "KRX 규격 5%")
put(16, "1포인트 가치", 1000000, "원", WON, "in", "액면 1억 ÷ 100")
put(17, "계약수 Q", 100, "계약", INT_, "in")
put(18, "포지션 방향", "선물매도+현물매수", "", None, "in", "선물매도+현물매수 / 선물매수+현물매도")
put(19, "방향계수", '=IF(B18="선물매도+현물매수",1,-1)', "", INT_)
put(20, "선물 진입가", None, "", "0.00", "in", "빈칸이면 현재가 사용")
put(21, "▶ 적용 진입가", '=IF(B20="",B25,B20)', "", "0.00", "key")
put(22, "거래비용", 13000, "원/계약", WON, "in", "선물수수료 + 채권 스프레드 + 슬리피지")

put(23, "대차수수료 (현물매도 시에만)", 0.0, "bp", BP, "in",
    "현물을 빌려서 파는 방향에서만 적용. 현물매수 방향에서는 무시된다 "
    "(산 채권을 대차로 빌려주는 것은 별도 영업이며, special 이점은 조달금리 B12 에 이미 반영됨)")

dv = DataValidation(type="list", formula1='"선물매도+현물매수,선물매수+현물매도"', allow_blank=True)
m.add_data_validation(dv)
dv.add(m["B18"])

# ── [2] 실시간 (인포맥스) ───────────────────────────────────
sec(24, "[2] 실시간 — 연합인포맥스   ◆ 붉은 셀이 IMDP 연결")
put(25, "선물 현재가", '=_xll.IMDP("FUT",$B$26,"선물호가_현재가")', "", "0.00", "live")
put(26, "선물 종목코드", "A6769000", "", None, "in", "최근월물 코드로 갱신")
put(27, "종목1 수익률", '=_xll.IMDP("BND",$C$34,"장내국채-매도수익률1")', "%", YLD, "live",
    "코드 인수를 셀참조로 씀. 오류 시 코드를 수식에 직접 입력")
put(28, "종목2 수익률", '=_xll.IMDP("BND",$D$34,"장내국채-매도수익률1")', "%", YLD, "live")
put(29, "인포맥스 이론가 (참고)", '=_xll.IMDP("FUT",$B$26,"이론가")', "", "0.00", "live",
    "B26과 같은 종목코드를 써야 비교가 유효")
m["A30"] = ("※ 종목1을 실시간으로 못 받으면 B27 을  =B28-2.6/100  처럼 스프레드 파생식으로 바꾸십시오 "
            "(원본 파일 방식). 단 스프레드는 수동 갱신이 필요합니다.")
as_text(m["A30"]).font = F_NOTE

put(31, "▶ 적용금리 (방향 반영)", "=IF($B$19=1,$B$12,$B$13-$B$23/100)", "%", YLD, "key",
    "현물매수 → 조달금리 B12 그대로.  현물매도 → 담보운용금리 B13 − 대차수수료 B23")

# ── [3] 바스켓 ──────────────────────────────────────────────
sec(32, "[3] 바스켓 (2종목)")
for col, h in zip("ABCDE", ["항목", "단위", "종목1", "종목2", "합계"]):
    c = m[f"{col}33"]
    c.value, c.font, c.fill, c.border = h, F_HDR, FILL_HDR, BOX

BOND = [("국고 25-5", "KR103502GE97", 2.625, dt.date(2025, 6, 10), dt.date(2035, 6, 10)),
        ("국고 25-11", "KR103503GG60", 3.250, dt.date(2025, 12, 10), dt.date(2035, 12, 10))]
EV = {"C": "B", "D": "AF"}      # 메인 열 → 엔진 스칼라 값 열


def brow(row, label, unit, fn, fmt=None, style="calc", total=None, note=""):
    as_text(m.cell(row=row, column=1, value=label)).font = F_LBL
    as_text(m.cell(row=row, column=2, value=unit)).font = F_NOTE
    for i, col in enumerate("CD"):
        c = m[f"{col}{row}"]
        c.value = fn(col, i) if callable(fn) else fn
        c.font = {"in": F_IN, "key": F_KEY, "live": F_LIVE}.get(style, F_CALC)
        fill = {"in": FILL_IN, "key": FILL_KEY, "live": FILL_LIVE}.get(style)
        if fill:
            c.fill = fill
        if fmt:
            c.number_format = fmt
        c.border = BOX
    if total:
        c = m[f"E{row}"]
        c.value, c.font, c.border = total, F_KEY, BOX
        if fmt:
            c.number_format = fmt
    if note:
        as_text(m.cell(row=row, column=6, value=note)).font = F_NOTE


brow(34, "종목코드", "", lambda c, i: BOND[i][1], None, "in")
brow(35, "종목명", "", lambda c, i: BOND[i][0], None, "in")
brow(36, "표면금리", "%", lambda c, i: BOND[i][2], YLD, "in")
brow(37, "발행일", "", lambda c, i: BOND[i][3], DATE, "in")
brow(38, "만기일", "", lambda c, i: BOND[i][4], DATE, "in", note="이표일을 여기서 6개월씩 역산")
brow(39, "가중치", "", 0.5, "0.000", "in", total="=SUM(C39:D39)",
     note="KRX 최종결제기준가격은 단순평균 → 2종목이면 0.5")
brow(40, "수익률 (실시간)", "%", lambda c, i: f"=B{27+i}", YLD, "live")
brow(41, "수익률 (소수)", "", lambda c, i: f"={c}40/100", "0.000000")
brow(42, "차기 이표일 (자동)", "", lambda c, i: f"=엔진!${EV[c]}$80", DATE,
     note="◀ 원본 파일은 여기가 하드코딩이라 스테일이었음")
brow(43, "직전 이표일 (자동)", "", lambda c, i: f"=엔진!${EV[c]}$81", DATE)
brow(44, "현물 단가 (더티)", "", lambda c, i: f"=엔진!${EV[c]}$84", PX, "key")
brow(45, "현물 BPV", "원/1bp/1억", lambda c, i: f"=엔진!${EV[c]}$87*$B$16", '#,##0.0')
brow(46, "수정듀레이션", "년", lambda c, i: f"=엔진!${EV[c]}$88", "0.000")
brow(47, "기간중 쿠폰 재투자FV", "", lambda c, i: f"=엔진!${EV[c]}$95", PX)
brow(48, "조달이자", "", lambda c, i: f"=엔진!${EV[c]}$97", PX)
brow(49, "선도 목표단가 (더티)", "", lambda c, i: f"=엔진!${EV[c]}$98", PX, "key",
     note="현물단가 + 조달이자 − 쿠폰FV")
brow(50, "▶ 선도 수익률", "%", lambda c, i: f"=엔진!${EV[c]}$109*100", YLD, "key")
brow(51, "수렴 잔차", "", lambda c, i: f"=엔진!${EV[c]}$111", "0.00E+00",
     note="0에 근접해야 정상")
brow(52, "선도 BPV", "포인트/1bp", lambda c, i: f"=엔진!${EV[c]}$108", "0.000000")
brow(53, "선도 BPV", "원/1bp/1억", lambda c, i: f"={c}52*$B$16", '#,##0.0')
brow(54, "캐리 (선도−현물)", "bp", lambda c, i: f"=({c}50-{c}40)*100", BP)
brow(55, "▶ 헤지액면 (계약 1개당)", "원",
     lambda c, i: f"=IFERROR({c}39*($B$60/{c}52)*$B$16*100,0)", WON, "key",
     total="=SUM(C55:D55)", note="w × (표준물BPV / 선도BPV) × 액면 1억")
brow(56, "헤지액면 총계 (×Q)", "원", lambda c, i: f"={c}55*$B$17", WON,
     total="=SUM(C56:D56)")
brow(57, "실제 보유액면 (입력)", "원", None, WON, "in", total="=SUM(C57:D57)")
brow(58, "▶ 적용 액면", "원", lambda c, i: f'=IF({c}57="",{c}56,{c}57)', WON, "key",
     total="=SUM(C58:D58)")

# ── [4] 이론가 / 베이시스 ───────────────────────────────────
sec(60 - 1, "[4] 이론가 · 베이시스")


def stdp(y):
    return (f"=($B$15*100/({y}))*(1-(1/(1+({y})/2))^($B$14*2))"
            f"+100*(1/(1+({y})/2))^($B$14*2)")


put(60, "▶ 표준물 BPV", "=(B62-B61)/2", "포인트/1bp", "0.000000", "key")
put(61, "  P(ȳ+1bp)", stdp("$B$65+0.0001"), "", "0.000000")
put(62, "  P(ȳ−1bp)", stdp("$B$65-0.0001"), "", "0.000000")
put(63, "표준물 BPV", "=B60*$B$16", "원/1bp/계약", '#,##0.0')
put(64, "바스켓 평균 선도수익률", "=SUMPRODUCT(C50:D50,C39:D39)/SUM(C39:D39)", "%", YLD, "key")
put(65, "  (소수)", "=B64/100", "", "0.00000000")
put(66, "▶ 이론 선물가격", stdp("$B$65"), "", PX, "key", "선도금리 기준 — 캐리 반영됨")
put(67, "시장 선물가격", "=B25", "", "0.00")
put(68, "▶ 베이시스", "=B67-B66", "포인트", PX, "key")
put(69, "▶ 베이시스", "=B68*$B$16", "원/계약", WON, "key")
put(70, "총 거래비용", "=$B$22", "원/계약", WON)
put(71, "▶ 순차익", "=ABS(B69)-B70", "원/계약", WON, "key")
put(72, "▶ 총 순차익", "=B71*$B$17", "원", WON, "key")
put(73, "권장 포지션",
    '=IF(B68>0,"선물 매도 + 바스켓 매수",IF(B68<0,"선물 매수 + 바스켓 매도","중립"))',
    "", None, "key")
put(74, "판정", '=IF(B71>0,"차익거래 가능","비용 미달 — 관망")', "", None, "key")

put(76, "[참고] 현물금리 평균", "=SUMPRODUCT(C40:D40,C39:D39)/SUM(C39:D39)", "%", YLD)
put(77, "[참고] 현물금리 기준 이론가", stdp("$B$76/100"), "", PX, "calc",
    "◀ 원본 파일 방식. 아래 차이만큼 캐리가 이미 들어있음")
put(78, "[참고] 위 둘의 차이 = 캐리", "=(B77-B66)*$B$16", "원/계약", WON, "calc",
    "이 값을 캐리로 또 더하면 이중계상")
put(79, "[참고] 인포맥스 이론가", "=B29", "", "0.00")
put(80, "[참고] 인포맥스 − 자체", '=IFERROR(B79-B66,"")', "포인트", PX)

# ── [5] 만기 손익 ───────────────────────────────────────────
sec(82, "[5] 만기 손익   ◆ 바스켓 1단위 = 선물 1계약")
put(83, "만기 수익률 shift", 0, "bp", BP, "in", "만기시점 바스켓 수익률 평행이동")
put(84, "만기 바스켓 평균수익률", "=B64+B83/100", "%", YLD)
put(85, "▶ 최종결제기준가격", stdp("$B$84/100"), "", PX, "key")
put(86, "선물 leg", "=($B$21-B85)*$B$19*$B$16*$B$17", "원", WON)
put(87, "현물 leg",
    "=$B$19*((엔진!$B$114-C49)/100*C58+(엔진!$AF$114-D49)/100*D58)", "원", WON,
    "선도 목표단가 대비 처분손익 (조달이자·쿠폰FV는 선도단가에 반영됨)")
put(88, "거래비용", "=-$B$22*$B$17", "원", WON)
put(89, "▶ 순 만기손익", "=B86+B87+B88", "원", WON, "key")
put(90, "▶ 바스켓 1단위당", "=IFERROR(B89/$B$17,0)", "원/계약", WON, "key")
put(91, "바스켓 투자금액", "=SUMPRODUCT(C44:D44,C58:D58)/100", "원", WON)
put(92, "연환산 수익률", '=IF(AND(B91>0,B11>0),B89/B91*365/B11*100,"")', "%", "0.000", "key")

# ── [6] 시나리오 ────────────────────────────────────────────
sec(94, "[6] 시나리오 — 헤지 유효성 (정밀 재할인)")
heads = ["shift(bp)", "평균수익률(%)", "최종결제가", "선물 leg(원)",
         "현물 leg(원)", "순 만기손익(원)", "계약당(원)"]
for i, h in enumerate(heads):
    c = m.cell(row=95, column=1 + i, value=h)
    c.font, c.fill, c.border = F_HDR, FILL_HDR, BOX
    c.alignment = Alignment(horizontal="center", wrap_text=True)

for i, sh in enumerate(SHIFTS):
    R = 96 + i
    m.cell(row=R, column=1, value=sh).number_format = BP
    m.cell(row=R, column=1).font = F_IN
    m.cell(row=R, column=2, value=f"=$B$64+$A{R}/100").number_format = YLD
    m.cell(row=R, column=3,
           value=stdp(f"$B{R}/100").replace("=", "", 1)).value = (
        "=" + stdp(f"$B{R}/100")[1:])
    m.cell(row=R, column=3).number_format = PX
    m.cell(row=R, column=4,
           value=f"=($B$21-$C{R})*$B$19*$B$16*$B$17").number_format = WON
    m.cell(row=R, column=5,
           value=f"=$B$19*((엔진!$B${115+i}-$C$49)/100*$C$58"
                 f"+(엔진!$AF${115+i}-$D$49)/100*$D$58)").number_format = WON
    m.cell(row=R, column=6, value=f"=$D{R}+$E{R}-$B$22*$B$17").number_format = WON
    m.cell(row=R, column=7, value=f"=IFERROR($F{R}/$B$17,0)").number_format = WON
    for cc in range(1, 8):
        m.cell(row=R, column=cc).border = BOX
        if cc > 1:
            m.cell(row=R, column=cc).font = F_CALC
    if sh == 0:
        for cc in range(1, 8):
            m.cell(row=R, column=cc).fill = FILL_KEY

R2 = 96 + len(SHIFTS) + 1
put(R2, "손익 최대", f"=MAX(F96:F{95+len(SHIFTS)})", "원", WON, "key")
put(R2 + 1, "손익 최소", f"=MIN(F96:F{95+len(SHIFTS)})", "원", WON, "key")
put(R2 + 2, "변동폭 (헤지 잔여리스크)", f"=B{R2}-B{R2+1}", "원", WON, "key")
m[f"A{R2+4}"] = ("※ BPV 헤지가 정확하면 순 만기손익이 shift와 무관하게 평평해야 합니다. "
                 "남는 변동은 표준물(5% 쿠폰)과 바스켓(저쿠폰)의 컨벡시티 차이이며 "
                 "BPV 헤지로는 제거되지 않는 구조적 잔여리스크입니다.")
as_text(m[f"A{R2+4}"]).font = F_NOTE
m[f"A{R2+5}"] = ("※ 현재 종목코드·수익률은 원본 파일에서 가져온 값입니다. "
                 "결제월이 바뀌면 [1] 결제월과 [3] 바스켓 종목을 KRX 공시 기준으로 갱신하십시오.")
as_text(m[f"A{R2+5}"]).font = F_NOTE
m.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════
# 엔진
# ═══════════════════════════════════════════════════════════
e = wb.create_sheet("엔진")
e["A1"] = "계산 엔진 — 이표 스케줄 / 한국 관행식 단가 / 선도수익률 (직접 수정 금지)"
e["A1"].font, e["A1"].fill = F_TITLE, FILL_TITLE
for c in range(1, 12):
    e.cell(row=1, column=c).fill = FILL_TITLE

HDR = ["k", "이표일", "CF", "현물포함", "현물idx", "PV@y", "PV@y+", "PV@y−",
       "선도포함", "선도idx", "PH@y0", "PH@y0+", "PH@y0−",
       "PH@y1", "PH@y1+", "PH@y1−", "PH@y2",
       "쿠폰FV", "PH@손익"] + [f"PH@S{i+1}" for i in range(len(SHIFTS))]

for bi, s in enumerate(BLOCKS):
    C = lambda o: get_column_letter(s + o)
    v = C(1)
    mc = "CD"[bi]
    e.cell(row=12, column=s, value=f"■ 종목{bi+1}").font = Font(
        name=FONT, size=10, bold=True, color="1F3864")
    e.column_dimensions[C(0)].width = 20
    e.column_dimensions[C(1)].width = 13
    for o in range(2, NCOL):
        e.column_dimensions[C(o)].width = 11

    for row, lab, f in [
        (2, "만기일", f"=메인!{mc}38"), (3, "발행일", f"=메인!{mc}37"),
        (4, "표면금리(소수)", f"=메인!{mc}36/100"), (5, "반기쿠폰", f"={v}4*100/2"),
        (6, "현물결제일 S0", "=메인!$B$10"), (7, "정산기준일 H", "=메인!$B$8"),
        (8, "현물수익률(소수)", f"=메인!{mc}41"), (9, "적용금리(소수)", "=메인!$B$31/100"),
        (10, "손익 shift(bp)", "=메인!$B$83"),
    ]:
        as_text(e.cell(row=row, column=s, value=lab)).font = F_LBL
        cc = e.cell(row=row, column=s + 1, value=f)
        cc.font, cc.border = F_CALC, BOX
        cc.number_format = DATE if row in (2, 3, 6, 7) else "0.000000"

    for o, h in enumerate(HDR):
        cc = e.cell(row=14, column=s + o, value=h)
        cc.font, cc.fill, cc.border = F_HDR, FILL_HDR, BOX
        cc.alignment = Alignment(horizontal="center", wrap_text=True)

    for i in range(NPER):
        R = SCH0 + i
        g = {o: C(o) for o in range(NCOL)}
        # 선도 할인율: y1(+shift) — 열 16 은 손익 shift, 17.. 은 시나리오
        shift_src = [f"${v}$10"] + [f"메인!$A${96+j}" for j in range(len(SHIFTS))]
        vals = {
            0: i + 1,
            1: f"=EDATE(${v}$2,-6*({NPER}-${g[0]}{R}))",
            2: f'=IF(AND({g[1]}{R}>${v}$3,{g[1]}{R}<=${v}$2),'
               f'IF({g[1]}{R}=${v}$2,${v}$5+100,${v}$5),0)',
            3: f'=IF(AND(${g[2]}{R}>0,{g[1]}{R}>${v}$6),1,0)',
            4: f'=IF({g[3]}{R}=1,SUM(${g[3]}${SCH0}:{g[3]}{R}),"")',
            5: f'=IF({g[3]}{R}=1,${g[2]}{R}/(1+${v}$8/2)^({g[4]}{R}-1),0)',
            6: f'=IF({g[3]}{R}=1,${g[2]}{R}/(1+(${v}$8+0.0001)/2)^({g[4]}{R}-1),0)',
            7: f'=IF({g[3]}{R}=1,${g[2]}{R}/(1+(${v}$8-0.0001)/2)^({g[4]}{R}-1),0)',
            8: f'=IF(AND(${g[2]}{R}>0,{g[1]}{R}>${v}$7),1,0)',
            9: f'=IF({g[8]}{R}=1,SUM(${g[8]}${SCH0}:{g[8]}{R}),"")',
            10: f'=IF({g[8]}{R}=1,${g[2]}{R}/(1+${v}$99/2)^({g[9]}{R}-1),0)',
            11: f'=IF({g[8]}{R}=1,${g[2]}{R}/(1+(${v}$99+0.0001)/2)^({g[9]}{R}-1),0)',
            12: f'=IF({g[8]}{R}=1,${g[2]}{R}/(1+(${v}$99-0.0001)/2)^({g[9]}{R}-1),0)',
            13: f'=IF({g[8]}{R}=1,${g[2]}{R}/(1+${v}$104/2)^({g[9]}{R}-1),0)',
            14: f'=IF({g[8]}{R}=1,${g[2]}{R}/(1+(${v}$104+0.0001)/2)^({g[9]}{R}-1),0)',
            15: f'=IF({g[8]}{R}=1,${g[2]}{R}/(1+(${v}$104-0.0001)/2)^({g[9]}{R}-1),0)',
            16: f'=IF({g[8]}{R}=1,${g[2]}{R}/(1+${v}$109/2)^({g[9]}{R}-1),0)',
            17: f'=IF(AND(${g[2]}{R}>0,{g[1]}{R}>${v}$6,{g[1]}{R}<=${v}$7),'
                f'${g[2]}{R}*(1+${v}$9*(${v}$7-{g[1]}{R})/365),0)',
        }
        for j, src in enumerate(shift_src):
            vals[18 + j] = (f'=IF({g[8]}{R}=1,${g[2]}{R}/'
                            f'(1+(${v}$109+{src}/10000)/2)^({g[9]}{R}-1),0)')
        for o, val in vals.items():
            cc = e.cell(row=R, column=s + o, value=val)
            cc.font = F_CALC
            cc.number_format = DATE if o == 1 else ("#,##0" if o in (0, 3, 4, 8, 9)
                                                    else "0.000000")

    rng = lambda o: f"{C(o)}{SCH0}:{C(o)}{SCH1}"
    scal = [
        (78, "── 현물 (S0) ──", None, None),
        (79, "잔존 이표횟수", f"=SUM({rng(3)})", INT_),
        (80, "차기 이표일", f"=INDEX({rng(1)},MATCH(1,{rng(3)},0))", DATE),
        (81, "직전 이표일", f"=EDATE({v}80,-6)", DATE),
        (82, "d", f"={v}80-{v}6", INT_),
        (83, "T", f"={v}80-{v}81", INT_),
        (84, "현물단가(더티)", f"=SUM({rng(5)})/(1+{v}8/2*{v}82/{v}83)", "0.000000"),
        (85, "P0 @+1bp", f"=SUM({rng(6)})/(1+({v}8+0.0001)/2*{v}82/{v}83)", "0.000000"),
        (86, "P0 @−1bp", f"=SUM({rng(7)})/(1+({v}8-0.0001)/2*{v}82/{v}83)", "0.000000"),
        (87, "현물 BPV", f"=({v}86-{v}85)/2", "0.000000"),
        (88, "수정듀레이션", f"={v}87*10000/{v}84", "0.0000"),
        (90, "── 선도 (H) ──", None, None),
        (91, "선도 차기이표일", f"=INDEX({rng(1)},MATCH(1,{rng(8)},0))", DATE),
        (92, "선도 직전이표일", f"=EDATE({v}91,-6)", DATE),
        (93, "d_H", f"={v}91-{v}7", INT_),
        (94, "T_H", f"={v}91-{v}92", INT_),
        (95, "쿠폰 재투자FV", f"=SUM({rng(17)})", "0.000000"),
        (96, "조달일수", f"={v}7-{v}6", INT_),
        (97, "조달이자", f"={v}84*{v}9*{v}96/365", "0.000000"),
        (98, "선도 목표단가", f"={v}84+{v}97-{v}95", "0.000000"),
        (99, "y0 (=현물수익률)", f"={v}8", "0.00000000"),
        (100, "P_H(y0)", f"=SUM({rng(10)})/(1+{v}99/2*{v}93/{v}94)", "0.000000"),
        (101, "P_H(y0+1bp)", f"=SUM({rng(11)})/(1+({v}99+0.0001)/2*{v}93/{v}94)", "0.000000"),
        (102, "P_H(y0−1bp)", f"=SUM({rng(12)})/(1+({v}99-0.0001)/2*{v}93/{v}94)", "0.000000"),
        (103, "|dP/dy|", f"=({v}102-{v}101)/0.0002", "0.000000"),
        (104, "y1 (뉴턴 1회)", f"={v}99+({v}100-{v}98)/{v}103", "0.00000000"),
        (105, "P_H(y1)", f"=SUM({rng(13)})/(1+{v}104/2*{v}93/{v}94)", "0.000000"),
        (106, "P_H(y1+1bp)", f"=SUM({rng(14)})/(1+({v}104+0.0001)/2*{v}93/{v}94)", "0.000000"),
        (107, "P_H(y1−1bp)", f"=SUM({rng(15)})/(1+({v}104-0.0001)/2*{v}93/{v}94)", "0.000000"),
        (108, "▶ 선도 BPV", f"=({v}107-{v}106)/2", "0.000000"),
        (109, "▶ 선도수익률 y2 (뉴턴 2회)", f"={v}104+({v}105-{v}98)/{v}103", "0.00000000"),
        (110, "P_H(y2)", f"=SUM({rng(16)})/(1+{v}109/2*{v}93/{v}94)", "0.000000"),
        (111, "수렴 잔차", f"={v}110-{v}98", "0.00E+00"),
        (113, "── 만기 재평가 ──", None, None),
    ]
    for row, lab, f, fmt in scal:
        c0 = as_text(e.cell(row=row, column=s, value=lab))
        if f is None:
            c0.font = Font(name=FONT, size=9, bold=True, color="1F3864")
            continue
        c0.font = F_LBL
        cc = e.cell(row=row, column=s + 1, value=f)
        cc.font, cc.number_format, cc.border = F_CALC, fmt, BOX

    # 만기 재평가: 113 = 손익 shift, 114.. = 시나리오
    labels = ["손익 shift"] + [f"시나리오 {sh:+d}bp" for sh in SHIFTS]
    srcs = [f"${v}$10"] + [f"메인!$A${96+j}" for j in range(len(SHIFTS))]
    for j, (lab, src) in enumerate(zip(labels, srcs)):
        row = 114 + j
        as_text(e.cell(row=row, column=s, value=f"만기단가 — {lab}")).font = F_LBL
        cc = e.cell(row=row, column=s + 1,
                    value=f"=SUM({rng(18+j)})/(1+({v}109+{src}/10000)/2*{v}93/{v}94)")
        cc.font, cc.number_format, cc.border = F_CALC, "0.000000", BOX

e.freeze_panes = "A15"
wb.save(OUT)
print("saved:", OUT)
