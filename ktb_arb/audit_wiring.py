# -*- coding: utf-8 -*-
"""
참조 배선 감사 (LibreOffice 불필요)

생성된 워크북의 모든 수식에서 셀 참조를 추출해
  1) 참조 대상 시트가 존재하는가
  2) 참조 대상 셀이 비어 있지 않은가 (오타로 인한 빈 셀 참조 = 무음 오류)
  3) 참조 대상 행의 라벨(A열)이 의도와 맞는가  ← 오프바이원 탐지
  4) 순환참조가 있는가
를 점검한다. 생성 스크립트의 행 번호 실수는 대부분 여기서 잡힌다.
"""
import re
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

XL = "/home/user/vmo/ktb_arb/국채선물_차익거래.xlsx"

REF = re.compile(
    r"(?:(?P<sheet>'[^']+'|[가-힣A-Za-z_][가-힣A-Za-z0-9_]*)!)?"
    r"\$?(?P<c1>[A-Z]{1,3})\$?(?P<r1>\d{1,5})"
    r"(?::\$?(?P<c2>[A-Z]{1,3})\$?(?P<r2>\d{1,5}))?"
)
FUNCS = {"IF", "AND", "OR", "SUM", "MAX", "MIN", "ABS", "ROUND", "INDEX", "MATCH",
         "SUMPRODUCT", "IFERROR", "AVERAGE", "EDATE", "WORKDAY", "TODAY", "NOW",
         "DATE", "WEEKDAY", "MOD", "VLOOKUP", "FALSE", "TRUE"}


def cells_of(rng, sheet, c1, r1, c2, r2):
    a, b = column_index_from_string(c1), column_index_from_string(c2 or c1)
    lo, hi = int(r1), int(r2 or r1)
    for cc in range(min(a, b), max(a, b) + 1):
        for rr in range(lo, hi + 1):
            yield (sheet, cc, rr)


def main():
    wf = load_workbook(XL)
    names = set(wf.sheetnames)

    # 각 시트의 라벨 열(대개 A열, 현금흐름은 블록 시작열)
    def label_of(sheet, row):
        ws = wf[sheet]
        for col in (1, 43, 85):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, str) and not v.startswith("="):
                return v
        return None

    problems = defaultdict(list)
    deps = {}
    total_formulas = 0

    for sn in wf.sheetnames:
        ws = wf[sn]
        for row in ws.iter_rows():
            for c in row:
                if not (isinstance(c.value, str) and c.value.startswith("=")):
                    continue
                total_formulas += 1
                body = c.value[1:]
                # 문자열 리터럴 제거
                body_nostr = re.sub(r'"[^"]*"', "", body)
                mine = set()
                for m in REF.finditer(body_nostr):
                    tok = m.group(0)
                    # 함수명 뒤에 붙은 것 배제
                    if m.group("sheet") is None:
                        pre = body_nostr[:m.start()]
                        if pre.endswith(tuple(f + "" for f in ("",))) and re.search(
                                r"[A-Z]$", pre):
                            continue
                    sheet = (m.group("sheet") or sn).strip("'")
                    if sheet not in names:
                        problems["없는 시트 참조"].append(f"{sn}!{c.coordinate}: {tok}")
                        continue
                    for ref in cells_of(tok, sheet, m.group("c1"), m.group("r1"),
                                        m.group("c2"), m.group("r2")):
                        mine.add(ref)
                deps[(sn, c.column, c.row)] = mine

                # 빈 셀 참조 점검 (범위 참조는 제외, 단일 참조만)
                singles = [m for m in REF.finditer(body_nostr) if m.group("c2") is None]
                for m in singles:
                    sheet = (m.group("sheet") or sn).strip("'")
                    if sheet not in names:
                        continue
                    tw = wf[sheet]
                    tgt = tw.cell(row=int(m.group("r1")),
                                  column=column_index_from_string(m.group("c1")))
                    if tgt.value is None and sheet != sn:
                        lbl = label_of(sheet, int(m.group("r1")))
                        problems["빈 셀 참조(타시트)"].append(
                            f"{sn}!{c.coordinate} → {sheet}!{m.group('c1')}{m.group('r1')}"
                            f"  [행라벨: {lbl}]  수식: {c.value[:60]}")

    # 순환참조 탐지
    color, cyc = {}, []

    def dfs(u, stack):
        color[u] = 1
        for v in deps.get(u, ()):
            if v not in deps:
                continue
            if color.get(v) == 1:
                cyc.append(" → ".join(f"{s}!{chr(64+cc) if cc<27 else cc}{rr}"
                                      for s, cc, rr in stack[stack.index(v):] + [v])
                           if v in stack else f"{v}")
            elif color.get(v, 0) == 0:
                dfs(v, stack + [v])
        color[u] = 2

    for u in list(deps):
        if color.get(u, 0) == 0:
            dfs(u, [u])

    print(f"총 수식 셀: {total_formulas:,}개")
    print(f"의존 그래프 노드: {len(deps):,}개\n")

    for k, v in problems.items():
        print(f"■ {k}: {len(v)}건")
        for line in v[:25]:
            print("   ", line)
        if len(v) > 25:
            print(f"    ... 외 {len(v)-25}건")
        print()

    print(f"■ 순환참조: {len(cyc)}건")
    for x in cyc[:10]:
        print("   ", x)

    # 핵심 참조 배선 스팟체크
    print("\n■ 핵심 참조 스팟체크 (참조 대상 행의 라벨 확인)")
    checks = [
        ("바스켓", "C25", "현금흐름", "B85", "단가(더티) P0"),
        ("바스켓", "C37", "현금흐름", "B111", "y2 (뉴턴 2회) ▶ 선도수익률"),
        ("바스켓", "D37", "현금흐름", "AR111", "y2 (뉴턴 2회) ▶ 선도수익률"),
        ("바스켓", "E37", "현금흐름", "CH111", "y2 (뉴턴 2회) ▶ 선도수익률"),
        ("바스켓", "C39", "현금흐름", "B116", "선도 BPV (포인트/1bp)"),
        ("바스켓", "C41", "현금흐름", "B117", "선도 2차미분 P''"),
        ("바스켓", "C36", "현금흐름", "B103", "선도 목표단가(더티)"),
        ("바스켓", "C34", "현금흐름", "B100", "기간중 쿠폰 재투자FV"),
        ("이론가", "B4", "바스켓", "C37", "▶ 선도 수익률"),
        ("이론가", "B18", None, None, None),
        ("손익", "B18", "바스켓", "F54", None),
        ("손익", "B41", "바스켓", "F51", None),
        ("손익", "B27", "현금흐름", "B121", "P_H @ y2+손익shift"),
        ("손익", "B28", "현금흐름", "AR121", "P_H @ y2+손익shift"),
        ("손익", "B29", "현금흐름", "CH121", "P_H @ y2+손익shift"),
        ("시나리오", "E4", "현금흐름", "B123", "P_H @ y2+시나리오1"),
        ("시나리오", "E14", "현금흐름", "B133", "P_H @ y2+시나리오11"),
        ("시나리오", "E24", "현금흐름", "B143", "P_H @ y2+시나리오21"),
    ]
    for sn, coord, tsheet, tcoord, expect_label in checks:
        f = wf[sn][coord].value
        line = f"  {sn}!{coord} = {f}"
        if tsheet and expect_label:
            r = int(re.search(r"\d+", tcoord).group())
            lbl = label_of(tsheet, r)
            ok = (lbl == expect_label)
            line += f"\n      → {tsheet}!{tcoord} 행라벨 = '{lbl}'  {'OK' if ok else '★불일치★'}"
        print(line)


main()
